import tkinter as tk
import ctypes
from tkinter import filedialog, messagebox
import os
import re
import openpyxl
import time
import copy
from tksheet import Sheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Make app DPI aware to prevent blurriness on high-DPI displays on Windows
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2) # PROCESS_PER_MONITOR_DPI_AWARE
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware() # Fallback for Windows 7/8
    except Exception:
        pass

# e-Way Bill Bulk Upload column names (46 fields)
COLUMNS = [
    "Supply Type *", "Sub Type *", "Doc Type *", "Doc No *", "Doc Date *", "Transaction Type *", 
    "From_OtherPartyName", "From_GSTIN *", "From_Address1", "From_Address2", "From_Place", 
    "Dispatch_Pin Code *", "Bill From_State *", "Dispatch From_State *", "To_OtherPartyName", 
    "To_GSTIN *", "To_Address1", "To_Address2", "To_Place", "Ship To_Pin Code *", 
    "Bill To_State *", "Ship To_State *", "Product", "Description", "HSN *", "Unit", "Qty", 
    "Assessable Value *", "Tax Rate (S+C+I+Cess+Cess Non Advol)", "CGST Amount", "SGST Amount", 
    "IGST Amount", "CESS Amount", "CESS Non Advol Amount", "Others", "Total Invoice Value *", 
    "Trans Mode", "Distance (Km) *", "Trans Name", "Trans ID", "Trans DocNo", "Trans Date", 
    "Vehicle No", "Vehicle Type", "Supply_type_desc", "Errors List"
]

# Curated Dark Theme Palette (Pure Black window with dark elements)
COLOR_BG = "#000000"          # Pure Black BG
COLOR_CARD = "#0A0B0C"        # Very Dark Gray Card BG
COLOR_INPUT_BG = "#16181A"    # Dark Input Field BG
COLOR_BORDER = "#212427"      # Dark Gray Border
COLOR_TEXT_PRI = "#F8F9FA"    # Soft White Text
COLOR_TEXT_SEC = "#ADB5BD"    # Light Muted Gray Text

# Accent Buttons (Sleek Dark Accent Colors)
COLOR_PRIMARY = "#0D6EFD"          # Premium Accent Blue
COLOR_PRIMARY_HOVER = "#0B5ED7"
COLOR_PRIMARY_TEXT = "#FFFFFF"     # White Text

COLOR_SUCCESS = "#198754"          # Accent Green
COLOR_SUCCESS_HOVER = "#157347"
COLOR_SUCCESS_TEXT = "#FFFFFF"     # White Text

COLOR_SECONDARY = "#212529"        # Accent Gray
COLOR_SECONDARY_HOVER = "#343A40"
COLOR_SECONDARY_TEXT = "#F8F9FA"   # Soft White Text

COLOR_DANGER = "#DC3545"           # Accent Red
COLOR_DANGER_HOVER = "#BB2D3B"
COLOR_DANGER_TEXT = "#FFFFFF"      # White Text

# --- Premium Custom UI Widgets ---

class ModernButton(tk.Canvas):
    def __init__(self, parent, text, command, bg=None, hover_bg=None, fg=None, font=("Segoe UI", 9, "bold"), radius=8, **kwargs):
        self.bg = bg or COLOR_PRIMARY
        self.hover_bg = hover_bg or COLOR_PRIMARY_HOVER
        
        # Match text color to button type
        if fg is None:
            if self.bg == COLOR_PRIMARY:
                self.fg = COLOR_PRIMARY_TEXT
            elif self.bg == COLOR_SUCCESS:
                self.fg = COLOR_SUCCESS_TEXT
            elif self.bg == COLOR_DANGER:
                self.fg = COLOR_DANGER_TEXT
            elif self.bg == COLOR_SECONDARY:
                self.fg = COLOR_SECONDARY_TEXT
            else:
                self.fg = COLOR_TEXT_PRI
        else:
            self.fg = fg
            
        # Measure text size
        lbl = tk.Label(parent, text=text, font=font)
        text_width = lbl.winfo_reqwidth()
        text_height = lbl.winfo_reqheight()
        lbl.destroy()
        
        width = text_width + 24
        height = text_height + 12
        
        try:
            parent_bg = parent.cget("bg")
        except:
            parent_bg = COLOR_BG
            
        super().__init__(parent, bg=parent_bg, bd=0, highlightthickness=0, width=width, height=height, cursor="hand2")
        self.text = text
        self.command = command
        self.font = font
        self.radius = radius
        self.width = width
        self.height = height
        
        self.draw_button(self.bg)
        
        self.bind("<Enter>", lambda e: self.draw_button(self.hover_bg))
        self.bind("<Leave>", lambda e: self.draw_button(self.bg))
        self.bind("<Button-1>", lambda e: self.on_click())
        
    def draw_button(self, color):
        self.delete("all")
        r = self.radius
        w = self.width
        h = self.height
        
        if r > h / 2:
            r = int(h / 2)
        if r > w / 2:
            r = int(w / 2)
            
        self.create_oval(0, 0, 2*r, 2*r, fill=color, outline=color)
        self.create_oval(w - 2*r, 0, w, 2*r, fill=color, outline=color)
        self.create_oval(0, h - 2*r, 2*r, h, fill=color, outline=color)
        self.create_oval(w - 2*r, h - 2*r, w, h, fill=color, outline=color)
        
        self.create_rectangle(r, 0, w - r, h, fill=color, outline=color)
        self.create_rectangle(0, r, w, h - r, fill=color, outline=color)
        
        self.create_text(w/2, h/2, text=self.text, fill=self.fg, font=self.font, justify="center")
        
    def on_click(self):
        if self.command:
            self.command()

class ModernEntry(tk.Frame):
    def __init__(self, parent, width=50, default_value="", **kwargs):
        super().__init__(
            parent,
            bg=COLOR_INPUT_BG,
            highlightthickness=1,
            highlightbackground="#3A3F46",
            highlightcolor=COLOR_PRIMARY
        )
        self.entry = tk.Entry(
            self,
            width=width,
            bg=COLOR_INPUT_BG,
            fg=COLOR_TEXT_PRI,
            insertbackground=COLOR_TEXT_PRI,
            bd=0,
            relief=tk.FLAT,
            font=("Segoe UI", 10),
            **kwargs
        )
        self.entry.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)
        if default_value:
            self.entry.insert(0, default_value)
            
        self.entry.bind("<FocusIn>", lambda e: self.config(highlightbackground=COLOR_PRIMARY))
        self.entry.bind("<FocusOut>", lambda e: self.config(highlightbackground="#3A3F46"))
            
    def get(self):
        return self.entry.get()
        
    def delete(self, first, last=tk.END):
        self.entry.delete(first, last)
        
    def insert(self, index, string):
        self.entry.insert(index, string)

from tksheet.sheet import Dropdown

class CustomDropdown(Dropdown):
    def __init__(self, *args, **kwargs):
        # Force flat outline with a 1-pixel border
        kwargs["outline_thickness"] = 1
        kwargs["outline_color"] = "#CCCCCC"  # Lite gray border
        super().__init__(*args, **kwargs)
        # Hide horizontal and vertical scrollbars for a clean modern flat look
        self.hide("x_scrollbar")
        self.hide("y_scrollbar")

    def reset(self, *args, **kwargs):
        # Force dark theme colors for the dropdown popups inside data cells
        kwargs["bg"] = "#16181A"
        kwargs["fg"] = "#F8F9FA"
        kwargs["select_bg"] = "#0D6EFD"
        kwargs["select_fg"] = "#FFFFFF"
        kwargs["outline_color"] = "#CCCCCC"  # Lite gray border
        super().reset(*args, **kwargs)
        self.hide("x_scrollbar")
        self.hide("y_scrollbar")

class EWayBillApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("E-Way Bill Preparation Tool")
        self.geometry("1280x700")
        self.configure(bg=COLOR_BG)

        # --- Dark Windows title bar ---
        # Makes the native Windows title bar dark on Windows 10/11.
        try:
            self.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            value = ctypes.c_int(1)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd,
                DWMWA_USE_IMMERSIVE_DARK_MODE,
                ctypes.byref(value),
                ctypes.sizeof(value)
            )
        except Exception:
            pass
        
        # Set Window Title Bar Icon
        try:
            import os
            import sys
            icon_file = "app_icon.ico"
            if os.path.exists(icon_file):
                self.iconbitmap(icon_file)
            elif hasattr(sys, "_MEIPASS"):
                bundled_icon = os.path.join(sys._MEIPASS, "app_icon.ico")
                if os.path.exists(bundled_icon):
                    self.iconbitmap(bundled_icon)
        except Exception:
            pass
        
        # Debounce tracking for hotkeys
        self.last_shortcut_time = 0.0
        
        # Undo stack: stores snapshots of grid data before each load (max 10 levels)
        self._undo_stack = []
        
        # Load persisted settings from config file
        self.load_settings()

        # --- Header Banner (all buttons in one bar) ---
        header_banner = tk.Frame(self, bg=COLOR_CARD, height=45, bd=0)
        header_banner.pack(fill=tk.X)
        header_banner.pack_propagate(False)
        
        # Settings Button stays on the RIGHT
        ModernButton(
            header_banner,
            text="⚙️ Settings",
            command=self.open_settings,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
        ).pack(side=tk.RIGHT, padx=(5, 12), pady=5)
        
        # 3 Action Buttons on the LEFT — same style as Settings
        ModernButton(
            header_banner,
            text="Load Data",
            command=self.load_challan,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
        ).pack(side=tk.LEFT, padx=(12, 5), pady=5)
        
        ModernButton(
            header_banner,
            text="Export to Excel",
            command=self.export_ewb,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
        ).pack(side=tk.LEFT, padx=5, pady=5)
        
        ModernButton(
            header_banner,
            text="Clear Grid",
            command=self.clear_grid,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
        ).pack(side=tk.LEFT, padx=5, pady=5)
        
        ModernButton(
            header_banner,
            text="Validate",
            command=self.validate_data,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
        ).pack(side=tk.LEFT, padx=5, pady=5)
        
        ModernButton(
            header_banner,
            text="Prepare JSON",
            command=self.prepare_json,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
        ).pack(side=tk.LEFT, padx=5, pady=5)
        
        # Lite navbar borderline (in between data header and parts text)
        accent_strip = tk.Frame(self, bg=COLOR_BORDER, height=1)
        accent_strip.pack(fill=tk.X)

        # --- Part A / Part B Section Labels ---
        parts_bar = tk.Frame(self, bg=COLOR_BG, bd=0)
        parts_bar.pack(fill=tk.X, padx=12, pady=(6, 6))
        
        # Part A Header Badge (Dark Blue)
        part_a_frame = tk.Frame(parts_bar, bg="#1A365D", bd=0, padx=10, pady=4)
        part_a_frame.pack(side=tk.LEFT)
        tk.Label(
            part_a_frame,
            text="Part A  —  Supply, Document & Tax Details  (Columns 1–36)",
            font=("Segoe UI", 9, "bold"),
            fg="#FFFFFF",
            bg="#1A365D"
        ).pack()
        
        # Part B Header Badge (Dark Green)
        part_b_frame = tk.Frame(parts_bar, bg="#064E3B", bd=0, padx=10, pady=4)
        part_b_frame.pack(side=tk.LEFT, padx=(10, 0))
        tk.Label(
            part_b_frame,
            text="Part B  —  Transport Details  (Columns 37–46)",
            font=("Segoe UI", 9, "bold"),
            fg="#FFFFFF",
            bg="#064E3B"
        ).pack()

        # Error Note Badge (Simulated Light Transparent Red)
        error_note_frame = tk.Frame(parts_bar, bg="#4C1D1D", bd=0, padx=10, pady=4)
        error_note_frame.pack(side=tk.LEFT, padx=(10, 0))
        tk.Label(
            error_note_frame,
            text="Please check for the errors in the last column of the sheet",
            font=("Segoe UI", 9, "bold"),
            fg="#FFFFFF",
            bg="#4C1D1D"
        ).pack()

        # --- Formula Bar ---
        formula_bar = tk.Frame(self, bg=COLOR_BG, bd=0)
        formula_bar.pack(fill=tk.X, padx=12, pady=(0, 6))
        
        # Border container for formula bar entry
        formula_border = tk.Frame(formula_bar, bg=COLOR_BORDER, padx=1, pady=1)
        formula_border.pack(fill=tk.X, expand=True)
        
        # Inner frame to hold the entry field and rocket emoji
        formula_inner = tk.Frame(formula_border, bg=COLOR_INPUT_BG)
        formula_inner.pack(fill=tk.X, expand=True)

        # Load rocket icon image with PIL
        self.rocket_img = None
        try:
            import os
            import sys
            from PIL import Image, ImageTk
            icon_path = "rocket_icon.png"
            if hasattr(sys, "_MEIPASS"):
                bundled_path = os.path.join(sys._MEIPASS, "rocket_icon.png")
                if os.path.exists(bundled_path):
                    icon_path = bundled_path
            
            if os.path.exists(icon_path):
                pil_img = Image.open(icon_path)
                self.rocket_img = ImageTk.PhotoImage(pil_img)
        except Exception:
            pass

        if self.rocket_img:
            rocket_label = tk.Label(
                formula_inner,
                image=self.rocket_img,
                bg=COLOR_INPUT_BG
            )
        else:
            rocket_label = tk.Label(
                formula_inner,
                text=" 🚀 ",
                font=("Segoe UI", 10),
                bg=COLOR_INPUT_BG,
                fg=COLOR_TEXT_PRI,
                padx=5
            )
        rocket_label.pack(side=tk.LEFT, padx=(10, 8))
        
        # Separator line between rocket and entry
        sep = tk.Frame(formula_inner, bg=COLOR_BORDER, width=1)
        sep.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        
        self.formula_entry = tk.Entry(
            formula_inner,
            bg=COLOR_INPUT_BG,
            fg=COLOR_TEXT_PRI,
            insertbackground=COLOR_TEXT_PRI,
            bd=0,
            relief=tk.FLAT,
            font=("Segoe UI", 10)
        )
        self.formula_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2, pady=4)
        
        self.formula_entry.bind("<Return>", self.commit_formula_change)
        self.formula_entry.bind("<FocusOut>", self.commit_formula_change)

        # --- Grid Area Card (Takes maximum space) ---
        grid_card = tk.Frame(self, bg=COLOR_CARD, bd=0)
        grid_card.pack(fill=tk.BOTH, expand=True, padx=0, pady=(0, 0))


        # Configure ttk style to use 'clam' theme to enable scrollbar customizations
        try:
            from tkinter import ttk
            style = ttk.Style()
            style.theme_use("clam")

            # Keep any native/ttk scrollbar elements fully dark too.
            dark_scroll = "#0A0B0C"
            dark_trough = "#0A0B0C"
            dark_thumb = "#2A2D31"
            style.configure(
                "Vertical.TScrollbar",
                background=dark_thumb,
                troughcolor=dark_trough,
                bordercolor=dark_scroll,
                arrowcolor=dark_thumb,
                relief="flat",
                borderwidth=0,
            )
            style.configure(
                "Horizontal.TScrollbar",
                background=dark_thumb,
                troughcolor=dark_trough,
                bordercolor=dark_scroll,
                arrowcolor=dark_thumb,
                relief="flat",
                borderwidth=0,
            )
            style.map(
                "Vertical.TScrollbar",
                background=[("active", "#3A3F46"), ("pressed", "#3A3F46")],
            )
            style.map(
                "Horizontal.TScrollbar",
                background=[("active", "#3A3F46"), ("pressed", "#3A3F46")],
            )
        except Exception:
            pass

        # Initialize tksheet Sheet widget with dark premium theme colors and smaller cells (9pt)
        self.sheet = Sheet(
            grid_card,
            header=COLUMNS,
            show_row_index=True,
            font=("Segoe UI", 9, "normal"),
            header_font=("Segoe UI", 9, "bold"),
            
            # Spreadsheet Colors
            bg="#D6DCE4",                  # Light blue-grey data bg
            fg="#212529",                  # Dark text for readable cells
            header_bg="#305496",           # Part A column header bg
            header_fg="#FFFFFF",           # White text for column headers
            index_bg="#262626",            # Grey row number (serial) bg
            index_fg="#FFFFFF",            # Dark text for row numbers
            top_left_bg="#305496",         # Matches column header bg
            grid_color=COLOR_BORDER,
            
            # Selection Colors
            selected_cells_bg="#B4C7E7",
            selected_rows_bg="#B4C7E7",
            selected_columns_bg="#B4C7E7",
            
            # Active Cell Colors
            active_cell_bg="#0D6EFD",
            active_cell_fg="#FFFFFF",
            resizing_line_color=COLOR_PRIMARY,
            
            # Sleek Dark Modern Scrollbars
            # Fully dark scrollbars
            vertical_scroll_background="#0A0B0C",
            horizontal_scroll_background="#0A0B0C",
            vertical_scroll_troughcolor="#0A0B0C",
            horizontal_scroll_troughcolor="#0A0B0C",
            vertical_scroll_bordercolor="#0A0B0C",
            horizontal_scroll_bordercolor="#0A0B0C",
            vertical_scroll_borderwidth=0,
            horizontal_scroll_borderwidth=0,
            vertical_scroll_relief="flat",
            horizontal_scroll_relief="flat",
            vertical_scroll_troughrelief="flat",
            horizontal_scroll_troughrelief="flat",
            vertical_scroll_not_active_bg="#16181A",
            horizontal_scroll_not_active_bg="#16181A",
            vertical_scroll_active_bg="#2A2D31",
            horizontal_scroll_active_bg="#2A2D31",
            vertical_scroll_pressed_bg="#3A3F46",
            horizontal_scroll_pressed_bg="#3A3F46",
            scrollbar_show_arrows=False
        )
        # Override the default dropdown class with our custom modern borderless one
        self.sheet._dropdown_cls = CustomDropdown
        
        # Standard alignment with 0 margin
        self.sheet.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.sheet.enable_bindings()
        
        # Bind extra sheet events to update the formula bar
        self.sheet.extra_bindings("cell_select", self.update_formula_bar)
        self.sheet.extra_bindings("drag_select_cells", self.update_formula_bar)
        self.sheet.extra_bindings("edit_cell", self.on_cell_edited)
        self.sheet.extra_bindings("deselect", self.update_formula_bar)
        
        # Bind row insertion events to re-apply cell dropdowns
        self.sheet.extra_bindings("end_insert_row", self.on_rows_inserted)
        self.sheet.extra_bindings("end_rc_insert_row", self.on_rows_inserted)
        
        # Highlight columns 37 to 44 (indices 36 to 43) with background #375623 and white text
        self.sheet.highlight_cells(cells=list(range(36, 44)), canvas="header", bg="#375623", fg="#FFFFFF")
        # Highlight Column 45 (index 44) with background #305496 and white text
        self.sheet.highlight_cells(cells=[44], canvas="header", bg="#305496", fg="#FFFFFF")
        # Highlight Column 46 (index 45) with background #DC3545 (red) and white text
        self.sheet.highlight_cells(cells=[45], canvas="header", bg="#DC3545", fg="#FFFFFF")
        
        # Narrow the serial number (row index) column width
        self.sheet.set_index_width(32)
        
        # Set slightly smaller default row and header heights
        self.sheet.row_height(height=20)
        self.sheet.set_header_height_pixels(22)
        
        # Initialize grid with clean empty slate
        self.sheet.set_sheet_data([])
        self.reapply_dropdowns()

        # --- Remove tksheet's Default Zoom Bindings ---
        for widget in (self.sheet.MT, self.sheet.RI, self.sheet.CH):
            try:
                widget.unbind("<Control-plus>")
                widget.unbind("<Control-equal>")
                widget.unbind("<Control-minus>")
            except Exception:
                pass

        # --- Keyboard Shortcut Bindings ---
        self.bind("<Control-plus>", self.shortcut_insert_row)
        self.bind("<Control-equal>", self.shortcut_insert_row)
        self.bind("<Control-Shift-equal>", self.shortcut_insert_row)
        self.bind("<Control-minus>", self.shortcut_delete_row)
        self.bind("<Control-z>", self.custom_undo)
        self.bind("<Control-Z>", self.custom_undo)
        
        # Bind to sheet components directly to intercept before default class bindings run
        for widget in (self.sheet.MT, self.sheet.RI, self.sheet.CH):
            try:
                widget.bind("<Control-z>", self.custom_undo)
                widget.bind("<Control-Z>", self.custom_undo)
            except Exception:
                pass

        # --- Bottom Status Bar ---
        self.status = tk.Label(
            self,
            text="Ready | Shortcuts: Ctrl++ add row, Ctrl+- delete row, Ctrl+Z to undo.",
            bd=0,
            relief=tk.FLAT,
            anchor=tk.W,
            bg=COLOR_CARD,
            fg=COLOR_TEXT_SEC,
            font=("Segoe UI", 9),
            padx=20,
            pady=6
        )
        self.status.pack(fill=tk.X, side=tk.BOTTOM)
        
        # Accent Border on top of Status Bar
        status_accent = tk.Frame(self, bg=COLOR_BORDER, height=1)
        status_accent.pack(fill=tk.X, side=tk.BOTTOM)

    # --- Formula Bar Event Handlers ---

    def on_cell_edited(self, event=None):
        self.update_formula_bar(event)
        import copy
        if event and isinstance(event, dict):
            r = event.get("row")
            c = event.get("column")
            if r is not None and c == 19:
                new_pincode = str(self.sheet.get_cell_data(r, c)).strip()
                if hasattr(self, "pincode_distances") and new_pincode in self.pincode_distances:
                    try:
                        custom_dist = int(float(self.pincode_distances[new_pincode]))
                        snapshot = copy.deepcopy(self.sheet.get_sheet_data())
                        self._undo_stack.append(snapshot)
                        if len(self._undo_stack) > 10:
                            self._undo_stack.pop(0)
                        self.sheet.set_cell_data(r, 37, custom_dist)
                        self.sheet.refresh()
                    except Exception:
                        pass

    def update_formula_bar(self, event=None):
        """Update formula bar entry box to show the content of the selected cell."""
        selected = self.sheet.get_currently_selected()
        if selected and hasattr(selected, "row") and hasattr(selected, "column"):
            r, c = selected.row, selected.column
            if r is not None and c is not None:
                val = self.sheet.get_cell_data(r, c)
                self.formula_entry.delete(0, tk.END)
                self.formula_entry.insert(0, str(val) if val is not None else "")
                return
        self.formula_entry.delete(0, tk.END)

    def commit_formula_change(self, event=None):
        """Sync editing changes from the formula bar back to the selected cell in the grid."""
        selected = self.sheet.get_currently_selected()
        if selected and hasattr(selected, "row") and hasattr(selected, "column"):
            r, c = selected.row, selected.column
            if r is not None and c is not None:
                new_val = self.formula_entry.get()
                current_val = self.sheet.get_cell_data(r, c)
                if str(new_val) != str(current_val):
                    # Capture undo snapshot
                    snapshot = copy.deepcopy(self.sheet.get_sheet_data())
                    self._undo_stack.append(snapshot)
                    if len(self._undo_stack) > 10:
                        self._undo_stack.pop(0)
                    
                    self.sheet.set_cell_data(r, c, new_val)
                    self.sheet.refresh()

    def get_default_row_data(self):
        import datetime
        today_str = datetime.date.today().strftime("%d/%m/%Y")
        row_data = ["" for _ in range(46)]
        row_data[0] = self.def_supply_type
        row_data[1] = self.def_sub_type
        row_data[2] = self.def_doc_type
        row_data[5] = self.def_transaction_type
        row_data[12] = self.def_bill_from_state
        row_data[13] = self.def_dispatch_from_state
        row_data[20] = self.def_bill_to_state
        row_data[21] = self.def_ship_to_state
        row_data[36] = self.def_trans_mode
        row_data[37] = self.default_distance
        row_data[38] = self.def_trans_name
        row_data[39] = self.def_trans_id
        row_data[41] = today_str
        row_data[42] = self.def_vehicle_no
        row_data[43] = self.def_vehicle_type
        row_data[44] = ""
        return row_data

    def on_rows_inserted(self, event=None):
        # Automatically populate new row(s) with defaults if empty
        if event and "added" in event and "rows" in event["added"] and "table" in event["added"]["rows"]:
            inserted_indices = event["added"]["rows"]["table"].keys()
            default_row = self.get_default_row_data()
            for r_idx in inserted_indices:
                for col_idx in [0, 1, 2, 5, 12, 13, 20, 21, 36, 37, 38, 39, 41, 42, 43]:
                    current_val = self.sheet.get_cell_data(r_idx, col_idx)
                    if current_val is None or str(current_val).strip() == "":
                        self.sheet.set_cell_data(r_idx, col_idx, default_row[col_idx])
        self.reapply_dropdowns()

    def reapply_dropdowns(self):
        """Re-apply/configure dropdowns on the tksheet columns."""
        dropdown_configs = {
            0: sorted(list(self.VALID_SUPPLY_TYPES)),
            1: sorted(list(self.VALID_SUB_TYPES)),
            2: sorted(list(self.VALID_DOC_TYPES)),
            5: sorted(list(self.VALID_TRANSACTION_TYPES)),
            12: sorted(list(self.VALID_STATES)),
            13: sorted(list(self.VALID_STATES)),
            20: sorted(list(self.VALID_STATES)),
            21: sorted(list(self.VALID_STATES)),
            36: sorted(list(self.VALID_TRANS_MODES)),
            43: sorted(list(self.VALID_VEHICLE_TYPES))
        }
        for col_idx, values in dropdown_configs.items():
            self.sheet.create_dropdown(
                r="all",
                c=col_idx,
                values=values,
                state="normal",  # Allows keyboard selection/searching and custom entry
                edit_data=False, # Prevent resetting cell data to first dropdown value
                redraw=False
            )
        self.sheet.refresh()

    # --- Keyboard Shortcut Handlers ---

    def custom_undo(self, event=None):
        """Ctrl+Z: Undo the last action (cell edit, row operation, or load operation)."""
        # If tksheet has its own edits to undo, use those first!
        tksheet_undo_stack = self.sheet.get_undo_stack()
        if tksheet_undo_stack and len(tksheet_undo_stack) > 0:
            try:
                self.sheet.undo()
                self.set_status("Last grid edit undone.")
                return "break"
            except Exception:
                pass

        # Otherwise, fall back to undoing the last file load
        if not self._undo_stack:
            self.set_status("Nothing to undo.")
            return "break"
        
        previous_state = self._undo_stack.pop()
        
        if previous_state:
            self.sheet.set_sheet_data(previous_state)
        else:
            # Undo back to empty grid
            self.sheet.set_sheet_data([["" for _ in range(46)]])
        
        self.reapply_dropdowns()
        self.sheet.set_all_row_heights(height=20)
        self.sheet.refresh()
        
        remaining = len(self._undo_stack)
        self.set_status(
            f"Undo successful. Grid restored to state before last load. "
            f"({remaining} more undo level{'s' if remaining != 1 else ''} available)"
        )
        return "break"

    def shortcut_insert_row(self, event=None):
        now = time.time()
        if now - self.last_shortcut_time < 0.15: # 150ms debounce window
            return "break"
        self.last_shortcut_time = now

        row_vals = self.get_default_row_data()
        selected = self.sheet.get_selected_rows()
        if selected:
            # Insert at the first selected row index
            idx = min(selected)
            self.sheet.insert_row(idx=idx, row=row_vals)
        else:
            curr = self.sheet.get_currently_selected()
            if curr:
                # Retrieve row index from namedtuple Selected
                idx = curr.row if hasattr(curr, "row") else curr[0]
                self.sheet.insert_row(idx=idx, row=row_vals)
            else:
                # Fallback to appending at the end of the sheet
                self.sheet.insert_row(row=row_vals)
                
        self.sheet.set_all_row_heights(height=20)
        self.sheet.refresh()
        self.set_status("Inserted row with default values (Ctrl++).")
        return "break"

    def shortcut_delete_row(self, event=None):
        now = time.time()
        if now - self.last_shortcut_time < 0.15:
            return "break"
        self.last_shortcut_time = now

        selected = self.sheet.get_selected_rows()
        if selected:
            # Delete in reverse order to keep indices valid
            for idx in sorted(selected, reverse=True):
                self.sheet.delete_row(idx)
            self.set_status(f"Deleted {len(selected)} selected row(s) (Ctrl+-).")
        else:
            curr = self.sheet.get_currently_selected()
            if curr:
                idx = curr.row if hasattr(curr, "row") else curr[0]
                self.sheet.delete_row(idx)
                self.set_status(f"Deleted row {idx + 1} (Ctrl+-).")
            else:
                self.set_status("No rows selected to delete.")
        self.sheet.refresh()
        return "break"

    # --- Settings Dialog Popup ---

    def load_settings(self):
        default_folder = r"F:\New pro 08-08-2026"
        if not os.path.exists(default_folder):
            default_folder = os.getcwd()
            
        self.settings_file = os.path.join(default_folder, "settings_config.json")
        
        # Standard defaults
        self.src_path = os.path.join(default_folder, "COCIW2608-0030.xlsx")
        self.tgt_path = os.path.join(default_folder, "summary.xlsx")
        self.default_distance = "100"
        self.default_unit = "UNITS"
        
        # Consignor details (Bill From profile)
        self.consignor_gstin = "32AAACH8025R2ZB"
        self.consignor_name = "HERBALIFE INTERNATIONAL INDIA PVT LTD"
        self.consignor_address = "C/O FIT 3PL WAREHOUSING PRIVATE LIMITED BUILDING"
        self.consignor_place = "ERNAKULAM"
        self.consignor_state = "Kerala"
        self.consignor_pincode = "683511"
        
        self.def_supply_type = "Outward"
        self.def_sub_type = "Supply"
        self.def_doc_type = "Delivery Challan"
        self.def_transaction_type = "Bill To-Ship To"
        self.def_bill_from_state = "Kerala"
        self.def_dispatch_from_state = "Kerala"
        self.def_bill_to_state = "Kerala"
        self.def_ship_to_state = "Kerala"
        self.def_trans_mode = "Road"
        self.def_vehicle_type = "Regular"
        self.def_trans_name = ""
        self.def_trans_id = ""
        self.def_vehicle_no = ""
        self.pincode_distances = {}
        
        if os.path.exists(self.settings_file):
            try:
                import json
                with open(self.settings_file, "r") as f:
                    cfg = json.load(f)
                    self.src_path = cfg.get("src_path", self.src_path)
                    self.tgt_path = cfg.get("tgt_path", self.tgt_path)
                    self.default_distance = cfg.get("default_distance", self.default_distance)
                    self.default_unit = cfg.get("default_unit", self.default_unit)
                    self.pincode_distances = cfg.get("pincode_distances", {})
                    
                    self.consignor_gstin = cfg.get("consignor_gstin", self.consignor_gstin)
                    self.consignor_name = cfg.get("consignor_name", self.consignor_name)
                    self.consignor_address = cfg.get("consignor_address", self.consignor_address)
                    self.consignor_place = cfg.get("consignor_place", self.consignor_place)
                    self.consignor_state = cfg.get("consignor_state", self.consignor_state)
                    self.consignor_pincode = cfg.get("consignor_pincode", self.consignor_pincode)
                    
                    self.def_supply_type = cfg.get("def_supply_type", self.def_supply_type)
                    self.def_sub_type = cfg.get("def_sub_type", self.def_sub_type)
                    self.def_doc_type = cfg.get("def_doc_type", self.def_doc_type)
                    self.def_transaction_type = cfg.get("def_transaction_type", self.def_transaction_type)
                    self.def_bill_from_state = cfg.get("def_bill_from_state", self.def_bill_from_state)
                    self.def_dispatch_from_state = cfg.get("def_dispatch_from_state", self.def_dispatch_from_state)
                    self.def_bill_to_state = cfg.get("def_bill_to_state", self.def_bill_to_state)
                    self.def_ship_to_state = cfg.get("def_ship_to_state", self.def_ship_to_state)
                    self.def_trans_mode = cfg.get("def_trans_mode", self.def_trans_mode)
                    self.def_vehicle_type = cfg.get("def_vehicle_type", self.def_vehicle_type)
                    self.def_trans_name = cfg.get("def_trans_name", self.def_trans_name)
                    self.def_trans_id = cfg.get("def_trans_id", self.def_trans_id)
                    self.def_vehicle_no = cfg.get("def_vehicle_no", self.def_vehicle_no)
            except Exception:
                pass

    def save_settings(self):
        try:
            import json
            cfg = {
                "src_path": self.src_path,
                "tgt_path": self.tgt_path,
                "default_distance": self.default_distance,
                "default_unit": self.default_unit,
                "pincode_distances": self.pincode_distances,
                "consignor_gstin": self.consignor_gstin,
                "consignor_name": self.consignor_name,
                "consignor_address": self.consignor_address,
                "consignor_place": self.consignor_place,
                "consignor_state": self.consignor_state,
                "consignor_pincode": self.consignor_pincode,
                
                "def_supply_type": self.def_supply_type,
                "def_sub_type": self.def_sub_type,
                "def_doc_type": self.def_doc_type,
                "def_transaction_type": self.def_transaction_type,
                "def_bill_from_state": self.def_bill_from_state,
                "def_dispatch_from_state": self.def_dispatch_from_state,
                "def_bill_to_state": self.def_bill_to_state,
                "def_ship_to_state": self.def_ship_to_state,
                "def_trans_mode": self.def_trans_mode,
                "def_vehicle_type": self.def_vehicle_type,
                "def_trans_name": self.def_trans_name,
                "def_trans_id": self.def_trans_id,
                "def_vehicle_no": self.def_vehicle_no,
            }
            with open(self.settings_file, "w") as f:
                json.dump(cfg, f, indent=4)
        except Exception:
            pass

    # --- Settings Dialog Popup ---

    def open_settings(self):
        from tkinter import ttk
        settings_win = tk.Toplevel(self)
        settings_win.attributes("-alpha", 0.0)  # Make invisible initially to prevent white flash
        settings_win.title("Configurations")
        
        # Set settings window icon
        try:
            import os
            import sys
            icon_file = "app_icon.ico"
            if os.path.exists(icon_file):
                settings_win.iconbitmap(icon_file)
            elif hasattr(sys, "_MEIPASS"):
                bundled_icon = os.path.join(sys._MEIPASS, "app_icon.ico")
                if os.path.exists(bundled_icon):
                    settings_win.iconbitmap(bundled_icon)
        except Exception:
            pass
        
        # Center settings window relative to main window
        # Center settings window relative to main window
        win_w = 820
        win_h = 580
        parent_x = self.winfo_x()
        parent_y = self.winfo_y()
        parent_w = self.winfo_width()
        parent_h = self.winfo_height()
        pos_x = parent_x + (parent_w - win_w) // 2
        pos_y = parent_y + (parent_h - win_h) // 2
        if pos_x < 0: pos_x = 0
        if pos_y < 0: pos_y = 0
        settings_win.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")
        
        settings_win.configure(bg=COLOR_BG)
        settings_win.resizable(False, False)
        settings_win.transient(self)
        settings_win.grab_set()
        
        # Make settings window title bar dark on Windows 10/11 (applied after transient/grab HWND change)
        try:
            settings_win.update()
            hwnd = ctypes.windll.user32.GetParent(settings_win.winfo_id())
            value = ctypes.c_int(1)
            # Try attributes 20 (recent Win 10/11) and 19 (older Win 10 versions)
            for attr in [20, 19]:
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd,
                    attr,
                    ctypes.byref(value),
                    ctypes.sizeof(value)
                )
        except Exception:
            pass
            
        settings_win.attributes("-alpha", 1.0)  # Make window fully visible

        # Apply dark modern flat styles to Combobox popups and widget
        settings_win.option_add("*TCombobox*Listbox.background", COLOR_INPUT_BG)
        settings_win.option_add("*TCombobox*Listbox.foreground", COLOR_TEXT_PRI)
        settings_win.option_add("*TCombobox*Listbox.selectBackground", COLOR_PRIMARY)
        settings_win.option_add("*TCombobox*Listbox.selectForeground", "#FFFFFF")
        settings_win.option_add("*TCombobox*Listbox.relief", "flat")
        settings_win.option_add("*TCombobox*Listbox.borderWidth", 0)

        style = ttk.Style(settings_win)
        try:
            style.theme_use('clam')
        except Exception:
            pass
        style.configure("TCombobox",
            fieldbackground=COLOR_INPUT_BG,
            background=COLOR_INPUT_BG,
            foreground=COLOR_TEXT_PRI,
            bordercolor="#3A3F46",
            darkcolor=COLOR_INPUT_BG,
            lightcolor=COLOR_INPUT_BG,
            arrowcolor="#FFFFFF",
            arrowsize=12,
            relief="flat",
            borderwidth=1
        )
        style.map("TCombobox",
            fieldbackground=[("readonly", COLOR_INPUT_BG)],
            background=[("readonly", COLOR_INPUT_BG)],
            foreground=[("readonly", COLOR_TEXT_PRI)],
            bordercolor=[("readonly", "#3A3F46")]
        )
        style.configure("Vertical.TScrollbar",
            gripcount=0,
            background="#3E444E",
            troughcolor="#1A1C1E",
            bordercolor="#1A1C1E",
            lightcolor="#3E444E",
            darkcolor="#3E444E",
            arrowcolor="#FFFFFF"
        )

        # Bottom Docked Buttons Frame (Fixed at bottom of window)
        btn_dock = tk.Frame(settings_win, bg=COLOR_BG, pady=10, padx=15)
        btn_dock.pack(side=tk.BOTTOM, fill=tk.X)

        # Outer Scrollable Container
        outer_container = tk.Frame(settings_win, bg=COLOR_BG)
        outer_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(outer_container, bg=COLOR_BG, highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(outer_container, orient="vertical", command=canvas.yview, style="Vertical.TScrollbar")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Scrollable inner frame (settings_card)
        settings_card = tk.Frame(canvas, bg=COLOR_CARD, bd=1, highlightbackground=COLOR_BORDER, highlightcolor=COLOR_BORDER, highlightthickness=1)
        canvas_frame_id = canvas.create_window((0, 0), window=settings_card, anchor="nw")

        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Match settings_card width to canvas width
            canvas.itemconfig(canvas_frame_id, width=canvas.winfo_width())

        canvas.bind("<Configure>", on_configure)
        settings_card.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        # Mousewheel binding
        settings_win.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        # Card Title 1
        tk.Label(
            settings_card,
            text="FILE PATHS & PROCESSING SETTINGS",
            font=("Segoe UI", 8, "bold"),
            fg=COLOR_PRIMARY,
            bg=COLOR_CARD,
            padx=15,
            pady=6
        ).grid(row=0, column=0, columnspan=3, sticky="w")

        # Target File Row
        tk.Label(settings_card, text="Target EWB Tool:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9, "bold")).grid(row=1, column=0, padx=(15, 10), pady=6, sticky="w")
        tgt_entry = ModernEntry(settings_card, width=70, default_value=self.tgt_path)
        tgt_entry.grid(row=1, column=1, padx=5, pady=6, sticky="ew")
        
        def browse_tgt():
            f = filedialog.askopenfilename(filetypes=[("Macro-Enabled Excel Files", "*.xlsm"), ("Excel Files", "*.xlsx"), ("All Files", "*.*")])
            if f:
                tgt_entry.delete(0, tk.END)
                tgt_entry.insert(0, f)
        ModernButton(settings_card, text="Browse...", command=browse_tgt, bg=COLOR_SECONDARY, hover_bg=COLOR_SECONDARY_HOVER).grid(row=1, column=2, padx=(10, 15), pady=6)

        # Horizontal Divider Line 0
        divider0 = tk.Frame(settings_card, bg=COLOR_BORDER, height=1)
        divider0.grid(row=2, column=0, columnspan=3, sticky="ew", pady=10)

        # Card Title 1.5 (Consignor details)
        tk.Label(
            settings_card,
            text="CONSIGNOR (BILL FROM) DETAILS",
            font=("Segoe UI", 8, "bold"),
            fg=COLOR_PRIMARY,
            bg=COLOR_CARD,
            padx=15,
            pady=4
        ).grid(row=3, column=0, columnspan=3, sticky="w")

        # Frame for Consignor details
        consignor_frame = tk.Frame(settings_card, bg=COLOR_CARD)
        consignor_frame.grid(row=4, column=0, columnspan=3, padx=15, pady=2, sticky="nsew")
        consignor_frame.columnconfigure(1, weight=1)
        consignor_frame.columnconfigure(3, weight=1)

        # Row 0: GSTIN, Legal Name
        tk.Label(consignor_frame, text="GSTIN/Trans ID *:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=0, padx=(10, 5), pady=3, sticky="w")
        consignor_gstin_entry = ModernEntry(consignor_frame, width=24, default_value=self.consignor_gstin)
        consignor_gstin_entry.grid(row=0, column=1, padx=(5, 30), pady=3, sticky="ew")

        tk.Label(consignor_frame, text="Legal Name:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=2, padx=(10, 5), pady=3, sticky="w")
        consignor_name_entry = ModernEntry(consignor_frame, width=24, default_value=self.consignor_name)
        consignor_name_entry.grid(row=0, column=3, padx=(5, 10), pady=3, sticky="ew")

        # Row 1: Address, Place
        tk.Label(consignor_frame, text="Address:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=0, padx=(10, 5), pady=3, sticky="w")
        consignor_address_entry = ModernEntry(consignor_frame, width=24, default_value=self.consignor_address)
        consignor_address_entry.grid(row=1, column=1, padx=(5, 30), pady=3, sticky="ew")

        tk.Label(consignor_frame, text="Place:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=2, padx=(10, 5), pady=3, sticky="w")
        consignor_place_entry = ModernEntry(consignor_frame, width=24, default_value=self.consignor_place)
        consignor_place_entry.grid(row=1, column=3, padx=(5, 10), pady=3, sticky="ew")

        states_list = sorted(list(self.VALID_STATES))

        # Row 2: State, Pin Code
        tk.Label(consignor_frame, text="State *:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=2, column=0, padx=(10, 5), pady=3, sticky="w")
        combo_consignor_state = ttk.Combobox(consignor_frame, values=states_list, state="readonly", width=22)
        combo_consignor_state.set(self.consignor_state)
        combo_consignor_state.grid(row=2, column=1, padx=(5, 30), pady=3, sticky="ew")

        tk.Label(consignor_frame, text="Pin Code *:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=2, column=2, padx=(10, 5), pady=3, sticky="w")
        consignor_pincode_entry = ModernEntry(consignor_frame, width=24, default_value=self.consignor_pincode)
        consignor_pincode_entry.grid(row=2, column=3, padx=(5, 10), pady=3, sticky="ew")

        # Horizontal Divider Line 1
        divider1 = tk.Frame(settings_card, bg=COLOR_BORDER, height=1)
        divider1.grid(row=5, column=0, columnspan=3, sticky="ew", pady=10)

        # Card Title 2 (Part A Defaults)
        tk.Label(
            settings_card,
            text="DEFAULT ROW VALUES - PART A",
            font=("Segoe UI", 8, "bold"),
            fg=COLOR_PRIMARY,
            bg=COLOR_CARD,
            padx=15,
            pady=4
        ).grid(row=6, column=0, columnspan=3, sticky="w")

        # Frame for Defaults grid to organize neatly (Part A)
        defaults_frame_a = tk.Frame(settings_card, bg=COLOR_CARD)
        defaults_frame_a.grid(row=7, column=0, columnspan=3, padx=15, pady=2, sticky="nsew")
        defaults_frame_a.columnconfigure(1, weight=1)
        defaults_frame_a.columnconfigure(3, weight=1)
        defaults_frame_a.columnconfigure(5, weight=1)

        # Row 0: Supply Type, Sub Type, Doc Type
        tk.Label(defaults_frame_a, text="Supply Type:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=0, padx=(10, 5), pady=3, sticky="w")
        combo_supply = ttk.Combobox(defaults_frame_a, values=sorted(list(self.VALID_SUPPLY_TYPES)), state="readonly", width=14)
        combo_supply.set(self.def_supply_type)
        combo_supply.grid(row=0, column=1, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_a, text="Sub Type:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=2, padx=(10, 5), pady=3, sticky="w")
        combo_sub = ttk.Combobox(defaults_frame_a, values=sorted(list(self.VALID_SUB_TYPES)), state="readonly", width=14)
        combo_sub.set(self.def_sub_type)
        combo_sub.grid(row=0, column=3, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_a, text="Doc Type:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=4, padx=(10, 5), pady=3, sticky="w")
        combo_doc = ttk.Combobox(defaults_frame_a, values=sorted(list(self.VALID_DOC_TYPES)), state="readonly", width=16)
        combo_doc.set(self.def_doc_type)
        combo_doc.grid(row=0, column=5, padx=(5, 10), pady=3, sticky="ew")

        # Row 1: Transaction Type, Bill From State, Dispatch From State
        tk.Label(defaults_frame_a, text="Transaction Type:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=0, padx=(10, 5), pady=3, sticky="w")
        combo_trans_type = ttk.Combobox(defaults_frame_a, values=sorted(list(self.VALID_TRANSACTION_TYPES)), state="readonly", width=14)
        combo_trans_type.set(self.def_transaction_type)
        combo_trans_type.grid(row=1, column=1, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_a, text="Bill From State:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=2, padx=(10, 5), pady=3, sticky="w")
        combo_bill_from = ttk.Combobox(defaults_frame_a, values=states_list, state="readonly", width=14)
        combo_bill_from.set(self.def_bill_from_state)
        combo_bill_from.grid(row=1, column=3, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_a, text="Dispatch From State:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=4, padx=(10, 5), pady=3, sticky="w")
        combo_dispatch_from = ttk.Combobox(defaults_frame_a, values=states_list, state="readonly", width=14)
        combo_dispatch_from.set(self.def_dispatch_from_state)
        combo_dispatch_from.grid(row=1, column=5, padx=(5, 10), pady=3, sticky="ew")

        # Row 2: Bill To State, Ship To State
        tk.Label(defaults_frame_a, text="Bill To State:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=2, column=0, padx=(10, 5), pady=3, sticky="w")
        combo_bill_to = ttk.Combobox(defaults_frame_a, values=states_list, state="readonly", width=14)
        combo_bill_to.set(self.def_bill_to_state)
        combo_bill_to.grid(row=2, column=1, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_a, text="Ship To State:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=2, column=2, padx=(10, 5), pady=3, sticky="w")
        combo_ship_to = ttk.Combobox(defaults_frame_a, values=states_list, state="readonly", width=14)
        combo_ship_to.set(self.def_ship_to_state)
        combo_ship_to.grid(row=2, column=3, padx=(5, 20), pady=3, sticky="ew")

        # Card Title 3 (Part B)
        tk.Label(
            settings_card,
            text="DEFAULT ROW VALUES - PART B",
            font=("Segoe UI", 8, "bold"),
            fg=COLOR_PRIMARY,
            bg=COLOR_CARD,
            padx=15,
            pady=4
        ).grid(row=8, column=0, columnspan=3, sticky="w")

        # Frame for Defaults grid (Part B)
        defaults_frame_b = tk.Frame(settings_card, bg=COLOR_CARD)
        defaults_frame_b.grid(row=9, column=0, columnspan=3, padx=15, pady=2, sticky="nsew")
        defaults_frame_b.columnconfigure(1, weight=1)
        defaults_frame_b.columnconfigure(3, weight=1)
        defaults_frame_b.columnconfigure(5, weight=1)

        # Row 0: Vehicle Type, Trans Mode, Trans Name
        tk.Label(defaults_frame_b, text="Vehicle Type:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=0, padx=(10, 5), pady=3, sticky="w")
        combo_vehicle = ttk.Combobox(defaults_frame_b, values=sorted(list(self.VALID_VEHICLE_TYPES)), state="readonly", width=14)
        combo_vehicle.set(self.def_vehicle_type)
        combo_vehicle.grid(row=0, column=1, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_b, text="Trans Mode:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=2, padx=(10, 5), pady=3, sticky="w")
        combo_trans_mode = ttk.Combobox(defaults_frame_b, values=sorted(list(self.VALID_TRANS_MODES)), state="readonly", width=14)
        combo_trans_mode.set(self.def_trans_mode)
        combo_trans_mode.grid(row=0, column=3, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_b, text="Trans Name:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=4, padx=(10, 5), pady=3, sticky="w")
        trans_name_entry = ModernEntry(defaults_frame_b, width=16, default_value=self.def_trans_name)
        trans_name_entry.grid(row=0, column=5, padx=(5, 10), pady=3, sticky="ew")

        # Row 1: Trans ID, Vehicle No
        tk.Label(defaults_frame_b, text="Trans ID:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=0, padx=(10, 5), pady=3, sticky="w")
        trans_id_entry = ModernEntry(defaults_frame_b, width=16, default_value=self.def_trans_id)
        trans_id_entry.grid(row=1, column=1, padx=(5, 20), pady=3, sticky="ew")

        tk.Label(defaults_frame_b, text="Vehicle No:", bg=COLOR_CARD, fg=COLOR_TEXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=2, padx=(10, 5), pady=3, sticky="w")
        vehicle_no_entry = ModernEntry(defaults_frame_b, width=16, default_value=self.def_vehicle_no)
        vehicle_no_entry.grid(row=1, column=3, padx=(5, 20), pady=3, sticky="ew")

        # Horizontal Divider Line 2
        divider2 = tk.Frame(settings_card, bg=COLOR_BORDER, height=1)
        divider2.grid(row=10, column=0, columnspan=3, sticky="ew", pady=10)

        # Card Title 4 (Pincode Distance Mapping)
        tk.Label(
            settings_card,
            text="PINCODE TO DISTANCE CUSTOM MAPPING (OPTIONAL)",
            font=("Segoe UI", 8, "bold"),
            fg=COLOR_PRIMARY,
            bg=COLOR_CARD,
            padx=15,
            pady=4
        ).grid(row=11, column=0, columnspan=3, sticky="w")

        # Frame for Pincode Distance Mapping
        mapping_frame = tk.Frame(settings_card, bg=COLOR_CARD)
        mapping_frame.grid(row=12, column=0, columnspan=3, padx=15, pady=2, sticky="nsew")
        tk.Label(
            mapping_frame,
            text="Enter custom distances (one per line, e.g., 682025: 120):",
            bg=COLOR_CARD,
            fg=COLOR_TEXT_SEC,
            font=("Segoe UI", 9)
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 5))

        # Initial formatted value from self.pincode_distances
        init_mapping_text = ""
        if hasattr(self, "pincode_distances") and self.pincode_distances:
            init_mapping_text = "\n".join(f"{pin}: {dist}" for pin, dist in self.pincode_distances.items())


        text_container = tk.Frame(mapping_frame, bg="#3A3F46", padx=1, pady=1)
        text_container.grid(row=1, column=0, columnspan=2, sticky="w")

        mapping_text = tk.Text(
            text_container,
            height=8,
            width=30,
            bg=COLOR_INPUT_BG,
            fg=COLOR_TEXT_PRI,
            insertbackground=COLOR_TEXT_PRI,
            bd=0,
            relief=tk.FLAT,
            font=("Segoe UI", 10)
        )
        mapping_text.pack(fill=tk.BOTH, expand=True)
        mapping_text.insert("1.0", init_mapping_text)

        # Focus bindings for the text widget
        mapping_text.bind("<FocusIn>", lambda e: text_container.config(bg=COLOR_PRIMARY))
        mapping_text.bind("<FocusOut>", lambda e: text_container.config(bg="#3A3F46"))

        def save():
            # self.src_path removed as load_challan prompts user directly
            self.tgt_path = tgt_entry.get().strip()
            
            self.consignor_gstin = consignor_gstin_entry.get().strip()
            self.consignor_name = consignor_name_entry.get().strip()
            self.consignor_address = consignor_address_entry.get().strip()
            self.consignor_place = consignor_place_entry.get().strip()
            self.consignor_state = combo_consignor_state.get()
            self.consignor_pincode = consignor_pincode_entry.get().strip()
            
            self.def_supply_type = combo_supply.get()
            self.def_sub_type = combo_sub.get()
            self.def_doc_type = combo_doc.get()
            self.def_transaction_type = combo_trans_type.get()
            self.def_bill_from_state = combo_bill_from.get()
            self.def_dispatch_from_state = combo_dispatch_from.get()
            self.def_bill_to_state = combo_bill_to.get()
            self.def_ship_to_state = combo_ship_to.get()
            
            self.def_trans_mode = combo_trans_mode.get()
            self.def_vehicle_type = combo_vehicle.get()
            self.def_trans_name = trans_name_entry.get().strip()
            self.def_trans_id = trans_id_entry.get().strip()
            self.def_vehicle_no = vehicle_no_entry.get().strip()

            # Parse custom pincode distance mapping
            raw_text = mapping_text.get("1.0", tk.END).strip().split("\n")

            new_mapping = {}
            for line in raw_text:
                line = line.strip()
                if not line:
                    continue
                for sep in [":", "=", "-"]:
                    if sep in line:
                        parts = line.split(sep, 1)
                        pin = parts[0].strip()
                        dist = parts[1].strip()
                        if pin.isdigit() and dist.replace(".", "", 1).isdigit():
                            new_mapping[pin] = dist
                        break
            self.pincode_distances = new_mapping
            
            self.save_settings()
            settings_win.destroy()
            self.set_status("Settings saved successfully.")
            
        # Dock buttons to btn_dock (always visible at bottom)
        ModernButton(
            btn_dock,
            text="Cancel",
            command=settings_win.destroy,
            bg=COLOR_SECONDARY,
            hover_bg=COLOR_SECONDARY_HOVER,
            width=12
        ).pack(side=tk.RIGHT, padx=5)

        ModernButton(
            btn_dock,
            text="Save Settings",
            command=save,
            bg=COLOR_PRIMARY,
            hover_bg=COLOR_PRIMARY_HOVER,
            width=14
        ).pack(side=tk.RIGHT, padx=5)

        settings_card.columnconfigure(1, weight=1)


    # --- Control Actions ---

    def clear_grid(self):
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all data in the grid?"):
            self.sheet.set_sheet_data([["" for _ in range(46)] for _ in range(1)])
            self.reapply_dropdowns()
            self.sheet.refresh()
            self.set_status("Grid cleared.")

    def set_status(self, text):
        self.status.config(text=f"Status: {text}")

    # --- Data Extraction and Mapping Logic ---

    def load_challan(self):
        from tkinter import filedialog
        import os
        
        filepath = filedialog.askopenfilename(
            title="Select Source Challan Excel File",
            filetypes=[("Excel Files", "*.xlsx;*.xls;*.xlsb"), ("All Files", "*.*")]
        )
        if not filepath:
            self.set_status("Load cancelled.")
            return

        if not os.path.exists(filepath):
            messagebox.showerror("Error", f"File path is invalid or does not exist:\n{filepath}")
            return

        if not hasattr(self, "loaded_filepaths"):
            self.loaded_filepaths = set()

        if filepath in self.loaded_filepaths:
            if not messagebox.askyesno("Duplicate File", f"The file '{os.path.basename(filepath)}' has already been loaded.\nDo you want to load it again?"):
                return
        
        try:
            self.set_status("Loading source challan file...")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "DC Document" not in wb.sheetnames:
                messagebox.showerror("Error", f"Sheet 'DC Document' not found in file:\n{filepath}")
                return
            
            ws = wb["DC Document"]
            
            # Read row values
            max_r = ws.max_row
            values = []
            for r in range(1, max_r + 1):
                values.append([ws.cell(row=r, column=c).value for c in range(1, 16)])

            # Extract Document No
            doc_no = ""
            h7_val = values[6][7] if len(values) > 6 and len(values[6]) > 7 else None
            if h7_val is not None and h7_val != "":
                doc_no = str(h7_val).strip()
            else:
                f6_val = values[5][5] if len(values) > 5 and len(values[5]) > 5 else None
                if f6_val is not None and f6_val != "":
                    doc_no = str(f6_val).replace("*", "").strip()

            # Check duplicate DC number in grid
            existing_rows = self.sheet.get_sheet_data()
            existing_docs = {str(row[3]).strip() for row in existing_rows if len(row) > 3 and str(row[3]).strip()}
            if doc_no and doc_no in existing_docs:
                if not messagebox.askyesno("Duplicate DC", f"Document Number '{doc_no}' is already loaded in the grid.\nDo you want to load it again?"):
                    return

            # Extract Document Date and normalize to DD/MM/YYYY format
            raw_date = values[5][2] if len(values) > 5 and len(values[5]) > 2 else None
            
            import datetime
            def parse_and_normalize_date(date_str):
                if not date_str:
                    return datetime.date.today().strftime("%d/%m/%Y")
                date_str = str(date_str).strip()
                if not date_str:
                    return datetime.date.today().strftime("%d/%m/%Y")
                formats = [
                    "%d/%m/%Y",  # 05/08/2026
                    "%d-%b-%Y",  # 05-Aug-2026
                    "%d-%B-%Y",  # 05-August-2026
                    "%Y-%m-%d",  # 2026-08-05
                    "%d-%m-%Y",  # 05-08-2026
                    "%d/%b/%Y",  # 05/Aug/2026
                    "%b %d, %Y", # Aug 05, 2026
                ]
                for fmt in formats:
                    try:
                        dt = datetime.datetime.strptime(date_str, fmt)
                        return dt.strftime("%d/%m/%Y")
                    except ValueError:
                        continue
                try:
                    serial = float(date_str)
                    dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=serial)
                    return dt.strftime("%d/%m/%Y")
                except ValueError:
                    pass
                return datetime.date.today().strftime("%d/%m/%Y")

            doc_date = parse_and_normalize_date(raw_date)

            # Extract Shipper Address & Pincode (No address splitting, no custom pincode formatting)
            from_addr_full = str(values[1][0] or "").strip() if len(values) > 1 else ""
            from_gstin = ""
            a4_val = str(values[3][0] or "") if len(values) > 3 else ""
            gstin_match = re.search(r"GSTIN:\s*([A-Z0-9]{15})", a4_val, re.IGNORECASE)
            if gstin_match:
                from_gstin = gstin_match.group(1).upper()
            else:
                for r in range(min(5, len(values))):
                    cell_val = str(values[r][0] or "")
                    m = re.search(r"\b\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}\b", cell_val, re.IGNORECASE)
                    if m:
                        from_gstin = m.group(0).upper()
                        break
            
            from_name = str(values[0][2] or "").strip() if len(values) > 0 and len(values[0]) > 2 else ""
            if not from_name:
                from_name = "ASUS TECHNOLOGY PVT. LTD."
            from_state = self.def_bill_from_state
            
            # Simple pincode: extract 5-6 digits from shipper address, do not change it
            from_pincode = ""
            pin_match = re.search(r"\b\d{5,6}\b", from_addr_full)
            if pin_match:
                from_pincode = pin_match.group(0)
            
            # Simple place: check if "ERNAKULAM" is in address
            from_place = "ERNAKULAM"
            if "ERNAKULAM" in from_addr_full.upper():
                from_place = "ERNAKULAM"
                
            def sanitize_address(addr):
                if not addr:
                    return ""
                cleaned = re.sub(r"[^a-zA-Z0-9\s\#\,\/\-\&]", "", str(addr))
                cleaned = re.sub(r"\s+", " ", cleaned).strip()
                return cleaned

            from_addr1_raw, from_addr2_raw = self.split_address(from_addr_full)
            from_addr1 = sanitize_address(from_addr1_raw)
            from_addr2 = sanitize_address(from_addr2_raw)

            # Extract Billed To details
            to_name = str(values[9][2] or "").strip() if len(values) > 9 and len(values[9]) > 2 else ""
            # Sanitize To_OtherPartyName: keep only letters and spaces
            to_name = re.sub(r"[^a-zA-Z\s]", "", to_name)
            to_name = re.sub(r"\s+", " ", to_name).strip()

            to_gstin = str(values[14][7] or "").strip() if len(values) > 14 and len(values[14]) > 7 else ""
            # Use Ship to Address for To_Address columns and split it
            to_addr_full = str(values[10][7] or "").strip() if len(values) > 10 and len(values[10]) > 7 else ""
            to_pincode = str(values[11][2] or "").strip() if len(values) > 11 and len(values[11]) > 2 else ""
            to_state = self.def_bill_to_state
            
            # Use Ship to City cell value directly
            to_place = str(values[12][7] or "").strip() if len(values) > 12 and len(values[12]) > 7 else ""
            if not to_place:
                to_place = "CALICUT"
                
            to_addr1_raw, to_addr2_raw = self.split_address(to_addr_full)
            to_addr1 = sanitize_address(to_addr1_raw)
            to_addr2 = sanitize_address(to_addr2_raw)

            # Extract Ship To details
            ship_to_pincode = str(values[11][7] or "").strip() if len(values) > 11 and len(values[11]) > 7 else ""
            ship_to_state = self.def_ship_to_state

            # Locate product table boundaries
            header_row_idx = -1
            total_row_idx = -1
            for r in range(10, len(values)):
                row_vals = values[r]
                if len(row_vals) > 1 and row_vals[0] == "No." and row_vals[1] == "Ebs No.":
                    header_row_idx = r
                if len(row_vals) > 5 and row_vals[5] == "TOTAL:":
                    total_row_idx = r
                    break
            
            if header_row_idx == -1 or total_row_idx == -1:
                messagebox.showerror("Error", "Could not locate the product table headers ('No.', 'Ebs No.') or the 'TOTAL:' row in the 'DC Document' sheet.")
                return

            # Extract distance and unit configs
            default_dist_val = int(self.default_distance) if self.default_distance.isdigit() else 100
            default_unit = self.default_unit or "UNITS"

            # Parse items and group by HSN
            grouped_items = {}
            for r in range(header_row_idx + 1, total_row_idx):
                row_vals = values[r]
                if not row_vals or len(row_vals) < 14 or row_vals[3] is None:
                    continue
                
                part_no = str(row_vals[3]).strip()
                part_desc = str(row_vals[4] or "").strip()
                hsn = str(row_vals[5] or "").strip()
                qty = float(row_vals[6] or 0)
                amount = float(row_vals[8] or 0)
                gst_pct_str = str(row_vals[9] or "").strip()
                cgst_amt = float(row_vals[10] or 0)
                sgst_amt = float(row_vals[11] or 0)
                igst_amt = float(row_vals[12] or 0)
                total_val = float(row_vals[13] or 0)

                # Parse GST rate percentage
                gst_pct = 0.0
                gst_match = re.search(r"(\d+(\.\d+)?)", gst_pct_str)
                if gst_match:
                    gst_pct = float(gst_match.group(1))

                if hsn not in grouped_items:
                    grouped_items[hsn] = {
                        "part_nos": [part_no],
                        "part_descs": [part_desc],
                        "qty": qty,
                        "amount": amount,
                        "gst_pct": gst_pct,
                        "cgst_amt": cgst_amt,
                        "sgst_amt": sgst_amt,
                        "igst_amt": igst_amt,
                        "total_val": total_val
                    }
                else:
                    item = grouped_items[hsn]
                    if part_no not in item["part_nos"]:
                        item["part_nos"].append(part_no)
                    if part_desc not in item["part_descs"]:
                        item["part_descs"].append(part_desc)
                    item["qty"] += qty
                    item["amount"] += amount
                    item["cgst_amt"] += cgst_amt
                    item["sgst_amt"] += sgst_amt
                    item["igst_amt"] += igst_amt
                    item["total_val"] += total_val
                    if gst_pct > item["gst_pct"]:
                        item["gst_pct"] = gst_pct

            # Calculate grand total invoice value (sum of all items in the invoice)
            grand_total_invoice_val = sum(item["total_val"] for item in grouped_items.values())

            # Construct E-Way bill rows from grouped items
            extracted_items = []
            for hsn, item in grouped_items.items():
                part_no = ", ".join(item["part_nos"])[:100]
                part_desc = ", ".join(item["part_descs"])[:250]
                part_desc = sanitize_address(part_desc)
                
                qty = item["qty"]
                amount = item["amount"]
                gst_pct = item["gst_pct"]
                cgst_amt = item["cgst_amt"]
                sgst_amt = item["sgst_amt"]
                igst_amt = item["igst_amt"]

                is_igst = igst_amt > 0 or (to_state != from_state and igst_amt == 0 and cgst_amt == 0 and sgst_amt == 0)
                if is_igst:
                    tax_rate_str = f"0+0+{gst_pct}+0+0"
                else:
                    half_rate = gst_pct / 2.0
                    tax_rate_str = f"{half_rate}+{half_rate}+0+0+0"

                row_data = ["" for _ in range(46)]
                row_data[0] = self.def_supply_type
                row_data[1] = self.def_sub_type
                row_data[2] = self.def_doc_type
                row_data[3] = doc_no
                row_data[4] = doc_date
                row_data[5] = self.def_transaction_type
                row_data[6] = from_name
                row_data[7] = from_gstin
                row_data[8] = from_addr1
                row_data[9] = from_addr2
                row_data[10] = from_place
                row_data[11] = int(from_pincode) if from_pincode.isdigit() else from_pincode
                row_data[12] = from_state
                row_data[13] = from_state
                row_data[14] = to_name
                row_data[15] = to_gstin
                row_data[16] = to_addr1
                row_data[17] = to_addr2
                row_data[18] = to_place
                row_data[19] = int(ship_to_pincode) if ship_to_pincode.isdigit() else ship_to_pincode
                row_data[20] = to_state
                row_data[21] = ship_to_state
                row_data[22] = ""
                row_data[23] = ""
                row_data[24] = int(hsn) if hsn.isdigit() else hsn
                row_data[25] = default_unit
                row_data[26] = ""
                row_data[27] = round(amount, 2)
                row_data[28] = tax_rate_str
                row_data[29] = round(cgst_amt, 2)
                row_data[30] = round(sgst_amt, 2)
                row_data[31] = round(igst_amt, 2)
                row_data[32] = 0.0
                row_data[33] = 0.0
                row_data[34] = 0.0
                row_data[35] = round(grand_total_invoice_val, 2)
                row_data[36] = self.def_trans_mode
                
                # Custom pincode distance mapping lookup
                pincode_clean = ship_to_pincode.strip()
                row_dist_val = ""
                if hasattr(self, "pincode_distances") and pincode_clean in self.pincode_distances:
                    try:
                        row_dist_val = int(float(self.pincode_distances[pincode_clean]))
                    except Exception:
                        pass
                row_data[37] = row_dist_val
                row_data[38] = self.def_trans_name
                row_data[39] = self.def_trans_id
                row_data[40] = ""
                row_data[41] = datetime.date.today().strftime("%d/%m/%Y")
                row_data[42] = self.def_vehicle_no
                row_data[43] = self.def_vehicle_type
                row_data[44] = ""
                row_data[45] = ""

                extracted_items.append(row_data)

            # Get current sheet data and append (filtering out completely empty rows)
            current_grid_data = self.sheet.get_sheet_data()
            current_grid_data = [row for row in current_grid_data if any(str(cell).strip() for cell in row)]
            
            # Save snapshot to undo stack BEFORE modifying grid (keep max 10 levels)
            snapshot = [row[:] for row in current_grid_data]
            self._undo_stack.append(snapshot)
            if len(self._undo_stack) > 10:
                self._undo_stack.pop(0)
            
            # Combine
            combined_grid_data = current_grid_data + extracted_items
            
            # Set values
            # Replace the grid data in one operation.
            # Do NOT recalculate every cell's size after each load: with
            # 46 columns and many rows that becomes a major bottleneck.
            self.sheet.set_sheet_data(combined_grid_data)
            self.reapply_dropdowns()
            self.sheet.set_all_row_heights(height=20)
            self.sheet.refresh()
            
            self.loaded_filepaths.add(filepath)
            messagebox.showinfo("Success", f"Successfully loaded and appended {len(extracted_items)} items from:\n{os.path.basename(filepath)}")
            self.set_status(f"Loaded {len(extracted_items)} items. Total grid items: {len(combined_grid_data)}.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load challan data:\n{e}")
            self.set_status("Error loading challan.")

    def export_ewb(self):
        target_path = self.tgt_path
        if not target_path:
            messagebox.showerror("Error", "Please configure the target Excel path in settings first.")
            return
        
        # Get grid data
        grid_data = self.sheet.get_sheet_data()
        # Filter out empty rows
        grid_data = [row for row in grid_data if any(str(cell).strip() for cell in row)]
        
        if not grid_data:
            messagebox.showwarning("Warning", "The grid is empty. Nothing to export.")
            return

        try:
            self.set_status("Writing data to spreadsheet...")
            
            # Check if file exists. If it is xlsm, we load with keep_vba=True.
            is_xlsm = target_path.lower().endswith(".xlsm")
            
            if os.path.exists(target_path):
                wb = openpyxl.load_workbook(target_path, keep_vba=is_xlsm)
            else:
                wb = openpyxl.Workbook()
                # Remove default active sheet if it's new
                if len(wb.sheetnames) > 0:
                    wb.remove(wb.active)
            
            sheet_name = "eWayBill"
            
            # If sheet exists, remove and recreate it to ensure absolutely clean default formatting
            if sheet_name in wb.sheetnames:
                idx = wb.sheetnames.index(sheet_name)
                wb.remove(wb[sheet_name])
                ws = wb.create_sheet(title=sheet_name, index=idx)
            else:
                ws = wb.create_sheet(title=sheet_name)
                
            # Set default styles on active sheet (standard Calibri 11, default heights)
            ws.views.sheetView[0].showGridLines = True
            
            # Write headers to row 1
            for c, col_name in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=c, value=col_name)
                cell.font = Font(name="Calibri", size=11, bold=True)
                
            # Write grid data starting from row 2
            for i, row_data in enumerate(grid_data):
                row_num = 2 + i
                ws.row_dimensions[row_num].height = 20 # Keep data row heights neat
                
                for j, val in enumerate(row_data):
                    cell_val = val
                    if isinstance(val, str) and val.strip() != "":
                        # check if it is numeric (but not things like GSTIN or Doc No which have leading zeros/letters)
                        if j in [11, 19, 24, 26, 27, 29, 30, 31, 32, 33, 34, 35, 37]: # Pincodes, HSN, Qty, Values, Distance
                            try:
                                if "." in val:
                                    cell_val = float(val)
                                else:
                                    cell_val = int(val)
                            except ValueError:
                                pass
                                
                    cell = ws.cell(row=row_num, column=1 + j, value=cell_val)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.number_format = "@"
                    
                    # Highlight Doc No (index 3) and Description (index 23)
                    if j == 3:
                        cell.fill = PatternFill(fill_type="solid", start_color="E8EAF6", end_color="E8EAF6")
                        cell.font = Font(name="Calibri", size=11, color="1A237E")
                    elif j == 23:
                        cell.fill = PatternFill(fill_type="solid", start_color="E0F2F1", end_color="E0F2F1")
                        cell.font = Font(name="Calibri", size=11, color="004D40")
                    else:
                        cell.font = Font(name="Calibri", size=11)

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10.0)

            wb.save(target_path)
            messagebox.showinfo("Export Successful", f"Successfully exported {len(grid_data)} rows of data to:\n{target_path}")
            self.set_status(f"Exported {len(grid_data)} rows successfully.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data:\n{e}")
            self.set_status("Error exporting data.")

    # --- Validation ---

    VALID_SUPPLY_TYPES = {"Inward", "Outward"}
    VALID_SUB_TYPES = {
        "Supply", "Import", "Export", "Job Work", "For Own Use", "Job work Returns",
        "Sales Return", "SKD/CKD/Lots", "Line Sales", "Recipient Not Known",
        "Exhibition or Fairs", "Others"
    }
    VALID_DOC_TYPES = {
        "Tax Invoice", "Bill of Supply", "Bill of Entry", "Delivery Challan",
        "Credit Note", "Others"
    }
    VALID_TRANSACTION_TYPES = {
        "Regular", "Bill To-Ship To", "Bill From-Dispatch From",
        "Combination of 2 and 3"
    }
    VALID_STATES = {
        "Andaman and Nicobar Islands", "Andhra Pradesh", "Arunachal Pradesh",
        "Assam", "Bihar", "Chandigarh", "Chhattisgarh", "Dadra and Nagar Haveli and Daman and Diu",
        "Delhi", "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jammu and Kashmir",
        "Jharkhand", "Karnataka", "Kerala", "Ladakh", "Lakshadweep",
        "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
        "Nagaland", "Odisha", "Other Territory", "Puducherry", "Punjab",
        "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura",
        "Uttar Pradesh", "Uttarakhand", "West Bengal"
    }
    VALID_TRANS_MODES = {"Road", "Rail", "Air", "Ship"}
    VALID_VEHICLE_TYPES = {"Regular", "Over Dimensional Cargo"}

    def show_validation_errors(self, total_errors, error_count_rows, lines_text):
        dialog = tk.Toplevel(self)
        dialog.attributes("-alpha", 0.0)
        dialog.title("Validation Failed")
        
        # Set validation errors dialog icon
        try:
            import os
            import sys
            icon_file = "app_icon.ico"
            if os.path.exists(icon_file):
                dialog.iconbitmap(icon_file)
            elif hasattr(sys, "_MEIPASS"):
                bundled_icon = os.path.join(sys._MEIPASS, "app_icon.ico")
                if os.path.exists(bundled_icon):
                    dialog.iconbitmap(bundled_icon)
        except Exception:
            pass
        
        # Geometry
        win_w = 720
        win_h = 540
        parent_x = self.winfo_x()
        parent_y = self.winfo_y()
        parent_w = self.winfo_width()
        parent_h = self.winfo_height()
        pos_x = parent_x + (parent_w - win_w) // 2
        pos_y = parent_y + (parent_h - win_h) // 2
        if pos_x < 0: pos_x = 0
        if pos_y < 0: pos_y = 0
        dialog.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")
        
        dialog.configure(bg=COLOR_BG)
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()
        
        # Dark Title Bar
        try:
            dialog.update()
            hwnd = ctypes.windll.user32.GetParent(dialog.winfo_id())
            value = ctypes.c_int(1)
            for attr in [20, 19]:
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd,
                    attr,
                    ctypes.byref(value),
                    ctypes.sizeof(value)
                )
        except Exception:
            pass
        dialog.attributes("-alpha", 1.0)
        
        # Main Container Card
        card = tk.Frame(dialog, bg=COLOR_CARD, bd=1, highlightbackground=COLOR_BORDER, highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Top Accent Alert Bar (Red for Validation Failed)
        accent_bar = tk.Frame(card, bg="#DC3545", height=4)
        accent_bar.pack(fill=tk.X, side=tk.TOP)
        
        # Header Area
        header_frame = tk.Frame(card, bg=COLOR_CARD)
        header_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        title_label = tk.Label(
            header_frame, 
            text="Validation Failed",
            font=("Segoe UI", 13, "bold"),
            fg="#DC3545",  # Soft Red
            bg=COLOR_CARD
        )
        title_label.pack(anchor="w")
        
        subtitle_label = tk.Label(
            header_frame, 
            text=f"Found {total_errors} error(s) across {error_count_rows} row(s). Correct the errors listed below:",
            font=("Segoe UI", 9),
            fg=COLOR_TEXT_SEC,
            bg=COLOR_CARD
        )
        subtitle_label.pack(anchor="w", pady=(4, 0))
        
        # Scrollable Text container
        text_container = tk.Frame(card, bg="#2E3033", padx=1, pady=1)
        text_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(5, 10))
        
        err_text_widget = tk.Text(
            text_container,
            bg=COLOR_INPUT_BG,
            fg=COLOR_TEXT_PRI,
            insertbackground=COLOR_TEXT_PRI,
            bd=0,
            relief=tk.FLAT,
            font=("Consolas", 10),
            padx=12,
            pady=12
        )
        err_text_widget.pack(fill=tk.BOTH, expand=True)
        
        # Populate text
        err_text_widget.insert("1.0", lines_text)
        err_text_widget.config(state="disabled") # Read-only
        
        # OK Button at bottom right
        btn_box = tk.Frame(card, bg=COLOR_CARD)
        btn_box.pack(fill=tk.X, side=tk.BOTTOM, padx=20, pady=(10, 20))
        
        ModernButton(
            btn_box, 
            text="Dismiss", 
            command=dialog.destroy, 
            bg=COLOR_PRIMARY, 
            hover_bg=COLOR_PRIMARY_HOVER,
            width=12
        ).pack(side=tk.RIGHT)


    def prepare_json(self):
        """Validates the grid data and generates the e-Way Bill JSON file."""
        grid_data = self.sheet.get_sheet_data()
        grid_data = [row for row in grid_data if any(str(cell).strip() for cell in row)]

        if not grid_data:
            messagebox.showinfo("Prepare JSON", "The grid is empty. Nothing to convert.")
            return

        # Run internal validation first
        self.set_status("Validating data for JSON generation...")
        total_errors = 0
        error_summary = []
        import datetime
        import re

        def normalize_date(date_str):
            if not date_str:
                return datetime.date.today().strftime("%d/%m/%Y")
            date_str = str(date_str).strip()
            if not date_str:
                return datetime.date.today().strftime("%d/%m/%Y")
            formats = [
                "%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%d-%m-%Y",
                "%d/%b/%Y", "%b %d, %Y"
            ]
            for fmt in formats:
                try:
                    dt = datetime.datetime.strptime(date_str, fmt)
                    return dt.strftime("%d/%m/%Y")
                except ValueError:
                    continue
            try:
                serial = float(date_str)
                dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=serial)
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                pass
            return datetime.date.today().strftime("%d/%m/%Y")

        def normalize_state(state_name):
            if not state_name:
                return None
            s_clean = str(state_name).strip().lower()
            for s in self.VALID_STATES:
                if s.lower() == s_clean:
                    return s
            return None

        def sanitize_address(addr):
            if not addr:
                return ""
            cleaned = re.sub(r"[^a-zA-Z0-9\s\#\,\/\-\&]", "", str(addr))
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            return cleaned

        for idx, row in enumerate(grid_data):
            row_num = idx + 1
            errors = []

            def val(col_idx):
                if col_idx < len(row):
                    return str(row[col_idx]).strip()
                return ""

            # supply type
            v = val(0)
            if not v or v not in self.VALID_SUPPLY_TYPES:
                errors.append(f"Supply Type '{v}' invalid")
            
            # sub type
            v = val(1)
            if not v or v not in self.VALID_SUB_TYPES:
                errors.append(f"Sub Type '{v}' invalid")
            
            # doc type
            v = val(2)
            if not v or v not in self.VALID_DOC_TYPES:
                errors.append(f"Doc Type '{v}' invalid")
                
            # doc no
            v = val(3)
            if not v:
                errors.append("Doc No is required")
            elif len(v) > 16:
                errors.append("Doc No too long (max 16)")
                
            # doc date
            v = val(4)
            norm_date = normalize_date(v)
            m = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', norm_date)
            if not m:
                errors.append(f"Doc Date '{v}' invalid")

            # to gstin
            v = val(15)
            if not v:
                errors.append("To GSTIN is required")

            # to place
            v = val(18)
            if not v:
                errors.append("To Place is required")

            # to pin code
            v = val(19)
            if not v or not v.isdigit() or len(v) != 6:
                errors.append("To Pin Code must be a 6-digit number")

            # HSN
            v = val(24)
            if not v or not v.isdigit():
                errors.append("HSN code is required and must be numeric")

            # Assessable value
            v = val(27)
            try:
                float(v)
            except ValueError:
                errors.append("Assessable Value must be a number")

            # Total Invoice Value
            v = val(35)
            try:
                float(v)
            except ValueError:
                errors.append("Total Invoice Value must be a number")

            if errors:
                total_errors += len(errors)
                error_summary.append((row_num, errors))

        if total_errors > 0:
            lines = []
            for row_num, errs in error_summary:
                lines.append(f"Row {row_num}: {len(errs)} error(s)")
                for e in errs:
                    lines.append(f"  • {e}")
                lines.append("")
            lines.append("Please resolve the validation errors before generating the JSON.")
            lines_text = "\n".join(lines)
            
            self.show_validation_errors(total_errors, len(error_summary), lines_text)
            self.set_status("JSON generation aborted due to validation errors.")
            return

        # Prompt for save path
        from tkinter import filedialog
        import json
        save_path = filedialog.asksaveasfilename(
            title="Save e-Way Bill JSON",
            defaultextension=".json",
            filetypes=[("JSON Files", "*.json")],
            initialfile="eway_bill_upload.json"
        )
        if not save_path:
            self.set_status("JSON preparation cancelled.")
            return

        # Helper state code converter
        STATE_CODE_MAP = {
            "Jammu and Kashmir": 1, "Himachal Pradesh": 2, "Punjab": 3, "Chandigarh": 4,
            "Uttarakhand": 5, "Haryana": 6, "Delhi": 7, "Rajasthan": 8, "Uttar Pradesh": 9,
            "Bihar": 10, "Sikkim": 11, "Arunachal Pradesh": 12, "Manipur": 13, "Mizoram": 14,
            "Tripura": 15, "Meghalaya": 17, "Assam": 18, "West Bengal": 19, "Jharkhand": 20,
            "Odisha": 21, "Chhattisgarh": 22, "Madhya Pradesh": 23, "Gujarat": 24,
            "Dadra and Nagar Haveli and Daman and Diu": 26, "Maharashtra": 27, "Karnataka": 29,
            "Goa": 30, "Lakshadweep": 31, "Kerala": 32, "Tamil Nadu": 33, "Puducherry": 34,
            "Andaman and Nicobar Islands": 35, "Telangana": 36, "Andhra Pradesh": 37,
            "Ladakh": 38, "Other Territory": 97
        }

        def get_state_code(state_name):
            if not state_name:
                return 32
            name_clean = str(state_name).strip().lower()
            for name, code in STATE_CODE_MAP.items():
                if name.lower() == name_clean:
                    return code
            return 32

        SUB_SUPPLY_TYPE_MAP = {
            "Supply": 1, "Import": 2, "Export": 3, "Job Work": 4, "For Own Use": 5,
            "Job Work Return": 6, "Sales Return": 7, "Others": 8, "SKD/CKD": 9,
            "Line Sales": 10, "Recipient Not Known": 11, "Exhibition or Fairs": 12
        }

        DOC_TYPE_MAP = {
            "Tax Invoice": "INV", "Bill of Supply": "BIL", "Bill of Entry": "BOE",
            "Delivery Challan": "CHL", "Others": "OTH"
        }

        TRANS_MODE_MAP = {
            "Road": 1, "Rail": 2, "Air": 3, "Ship": 4
        }

        VEHICLE_TYPE_MAP = {
            "Regular": "R", "Over Dimensional Cargo": "O"
        }

        def to_int_safe(val, default=0):
            try:
                if not val or str(val).strip() == "":
                    return default
                return int(float(str(val).strip()))
            except Exception:
                return default

        def to_float_safe(val, default=0.0):
            try:
                if not val or str(val).strip() == "":
                    return default
                return float(str(val).strip())
            except Exception:
                return default

        # Group rows by Doc No (column 3)
        from collections import defaultdict
        grouped_invoices = defaultdict(list)
        for row in grid_data:
            doc_no = str(row[3]).strip()
            grouped_invoices[doc_no].append(row)

        bill_lists = []
        for doc_no, rows in grouped_invoices.items():
            first_row = rows[0]
            
            # Parsing consignor info from settings
            c_gstin = getattr(self, "consignor_gstin", "32AAACH8025R2ZB")
            c_name = getattr(self, "consignor_name", "HERBALIFE INTERNATIONAL INDIA PVT LTD")
            c_address = getattr(self, "consignor_address", "C/O FIT 3PL WAREHOUSING PRIVATE LIMITED BUILDING")
            c_place = getattr(self, "consignor_place", "ERNAKULAM")
            c_state = getattr(self, "consignor_state", "Kerala")
            c_pincode = getattr(self, "consignor_pincode", "683511")
            
            # Map address split
            c_addr1, c_addr2 = self.split_address(c_address)

            # Map fromStateCode
            c_state_code = get_state_code(c_state)

            # Read from_gstin override from first row if present
            from_gstin = str(first_row[7]).strip() or c_gstin
            from_trd_name = str(first_row[6]).strip() or c_name
            from_addr1 = str(first_row[8]).strip() or c_addr1
            from_addr2 = str(first_row[9]).strip() or c_addr2
            from_place = str(first_row[10]).strip() or c_place
            from_pincode = to_int_safe(first_row[11]) or to_int_safe(c_pincode)
            from_state_code = get_state_code(first_row[12]) if str(first_row[12]).strip() else c_state_code
            act_from_state_code = get_state_code(first_row[13]) if str(first_row[13]).strip() else from_state_code

            # Read to party details
            to_gstin = str(first_row[15]).strip()
            to_trd_name = str(first_row[14]).strip()
            to_addr1 = str(first_row[16]).strip()
            to_addr2 = str(first_row[17]).strip()
            to_place = str(first_row[18]).strip()
            to_pincode = to_int_safe(first_row[19])
            to_state_code = get_state_code(first_row[20])
            act_to_state_code = get_state_code(first_row[21])

            # Calculate totals for items
            total_taxable_value = 0.0
            cgst_value = 0.0
            sgst_value = 0.0
            igst_value = 0.0
            cess_value = 0.0
            cess_non_advol = 0.0
            oth_value = 0.0

            # Sort items by HSN code (column 24) in ascending order
            sorted_rows = sorted(rows, key=lambda r: str(r[24]).strip())

            item_list = []
            for item_idx, r in enumerate(sorted_rows):
                taxable_amount = to_float_safe(r[27])
                cgst_amt = to_float_safe(r[29])
                sgst_amt = to_float_safe(r[30])
                igst_amt = to_float_safe(r[31])
                cess_amt = to_float_safe(r[32])
                c_non_advol = to_float_safe(r[33])
                oth_amt = to_float_safe(r[34])

                total_taxable_value += taxable_amount
                cgst_value += cgst_amt
                sgst_value += sgst_amt
                igst_value += igst_amt
                cess_value += cess_amt
                cess_non_advol += c_non_advol
                oth_value += oth_amt

                # Parse rates from column 28
                rate_str = str(r[28]).strip()
                parts = rate_str.split("+")
                cgst_rate = 0.0
                sgst_rate = 0.0
                igst_rate = 0.0
                cess_rate = 0.0
                if len(parts) >= 3:
                    try:
                        cgst_rate = float(parts[0])
                        sgst_rate = float(parts[1])
                        igst_rate = float(parts[2])
                        if len(parts) >= 4:
                            cess_rate = float(parts[3])
                    except Exception:
                        pass

                qty_val = to_float_safe(r[26])
                quantity_clean = int(qty_val) if qty_val.is_integer() else qty_val
                
                # Unit normalization to standard E-Way Bill 3-character codes
                unit_clean = str(r[25]).strip().upper()
                def_unit_map = {
                    "UNITS": "UNT", "UNIT": "UNT", "BAGS": "BAG", "BOXES": "BOX", "KGS": "KGS"
                }
                qty_unit = def_unit_map.get(unit_clean, unit_clean)

                item_list.append({
                    "itemNo": item_idx + 1,
                    "productName": str(r[22]).strip(),
                    "productDesc": str(r[23]).strip(),
                    "hsnCode": str(r[24]).strip(),
                    "quantity": quantity_clean,
                    "qtyUnit": qty_unit,
                    "taxableAmount": round(taxable_amount, 2),
                    "sgstRate": float(sgst_rate),
                    "cgstRate": float(cgst_rate),
                    "igstRate": float(igst_rate),
                    "cessRate": float(cess_rate),
                    "cessNonAdvol": round(c_non_advol, 2)
                })

            # Main HSN Code
            main_hsn = to_int_safe(first_row[24])

            # Transport info
            t_mode = TRANS_MODE_MAP.get(str(first_row[36]).strip(), 1)
            t_distance = to_int_safe(first_row[37])
            t_name = str(first_row[38]).strip()
            t_id = str(first_row[39]).strip()
            t_doc_no = str(first_row[40]).strip()
            t_doc_date = str(first_row[41]).strip()
            v_no = str(first_row[42]).strip()
            v_type = VEHICLE_TYPE_MAP.get(str(first_row[43]).strip(), "R")

            # Document date format normalization
            doc_date_val = normalize_date(str(first_row[4]))

            # Grand invoice value
            tot_inv_val = to_float_safe(first_row[35])

            # transType mapping for Transaction Type *
            TRANS_TYPE_MAP_JSON = {
                "Regular": 1, "Bill To-Ship To": 2, "Bill From-Dispatch From": 3,
                "Combination of 2 and 3": 4
            }
            trans_type_val = TRANS_TYPE_MAP_JSON.get(str(first_row[5]).strip(), 1)

            # subSupplyDesc mapping
            sub_supply_desc_val = str(first_row[1]).strip() if str(first_row[1]).strip() == "Others" else ""

            bill_lists.append({
                "userGtin": from_gstin,
                "supplyType": str(first_row[0]).strip()[0].upper() if str(first_row[0]).strip() else "O",
                "subSupplyType": SUB_SUPPLY_TYPE_MAP.get(str(first_row[1]).strip(), 1),
                "subSupplyDesc": sub_supply_desc_val,
                "docType": DOC_TYPE_MAP.get(str(first_row[2]).strip(), "INV"),
                "docNo": doc_no,
                "docDate": doc_date_val,
                "transType": trans_type_val,
                "fromGstin": from_gstin,
                "fromTrdName": from_trd_name,
                "fromAddr1": from_addr1,
                "fromAddr2": from_addr2,
                "fromPlace": from_place,
                "fromPincode": from_pincode,
                "fromStateCode": from_state_code,
                "actualFromStateCode": act_from_state_code,
                "toGstin": to_gstin,
                "toTrdName": to_trd_name,
                "toAddr1": to_addr1,
                "toAddr2": to_addr2,
                "toPlace": to_place,
                "toPincode": to_pincode,
                "toStateCode": to_state_code,
                "actualToStateCode": act_to_state_code,
                "totalValue": round(total_taxable_value, 2),
                "cgstValue": round(cgst_value, 2),
                "sgstValue": round(sgst_value, 2),
                "igstValue": round(igst_value, 2),
                "cessValue": round(cess_value, 2),
                "TotNonAdvolVal": round(cess_non_advol, 2),
                "OthValue": round(oth_value, 2),
                "totInvValue": round(tot_inv_val, 2),
                "transMode": t_mode,
                "transDistance": t_distance,
                "transporterName": t_name,
                "transporterId": t_id,
                "transDocNo": t_doc_no,
                "transDocDate": t_doc_date,
                "vehicleNo": v_no,
                "vehicleType": v_type,
                "mainHsnCode": main_hsn,
                "itemList": item_list
            })

        output_json = {
            "version": "1.0.0621",
            "billLists": bill_lists
        }

        try:
            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(output_json, f, indent=4, separators=(',', ':'))
            self.set_status(f"JSON prepared successfully: {save_path}")
            messagebox.showinfo("Success ✅", f"e-Way Bill JSON generated successfully at:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error ❌", f"Failed to save JSON file:\n{str(e)}")
            self.set_status("Error saving JSON.")

    def validate_data(self):
        """Validate all grid data against e-Way Bill bulk upload rules."""
        grid_data = self.sheet.get_sheet_data()
        grid_data = [row for row in grid_data if any(str(cell).strip() for cell in row)]

        if not grid_data:
            messagebox.showinfo("Validate", "The grid is empty. Nothing to validate.")
            return

        self.set_status("Validating data...")
        total_errors = 0
        error_summary = []  # list of (row_num, list_of_errors)

        import datetime

        def normalize_date(date_str):
            if not date_str:
                return datetime.date.today().strftime("%d/%m/%Y")
            
            date_str = str(date_str).strip()
            if not date_str:
                return datetime.date.today().strftime("%d/%m/%Y")
            
            formats = [
                "%d/%m/%Y",  # 05/08/2026
                "%d-%b-%Y",  # 05-Aug-2026
                "%d-%B-%Y",  # 05-August-2026
                "%Y-%m-%d",  # 2026-08-05
                "%d-%m-%Y",  # 05-08-2026
                "%d/%b/%Y",  # 05/Aug/2026
                "%b %d, %Y", # Aug 05, 2026
            ]
            
            for fmt in formats:
                try:
                    dt = datetime.datetime.strptime(date_str, fmt)
                    return dt.strftime("%d/%m/%Y")
                except ValueError:
                    continue
                    
            # Try parsing Excel float serial date (e.g. 46237.0)
            try:
                serial = float(date_str)
                dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=serial)
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                pass
                
            return datetime.date.today().strftime("%d/%m/%Y")

        def normalize_state(state_name):
            if not state_name:
                return None
            s_clean = str(state_name).strip().lower()
            for s in self.VALID_STATES:
                if s.lower() == s_clean:
                    return s
            return None

        def sanitize_address(addr):
            if not addr:
                return ""
            cleaned = re.sub(r"[^a-zA-Z0-9\s\#\,\/\-\&]", "", str(addr))
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            return cleaned

        for idx, row in enumerate(grid_data):
            row_num = idx + 1
            errors = []

            def val(col_idx):
                if col_idx < len(row):
                    return str(row[col_idx]).strip()
                return ""

            # --- Part A Validations ---

            # [0] Supply Type * (Required, must be Inward/Outward)
            v = val(0)
            if not v:
                errors.append("Supply Type is required")
            elif v not in self.VALID_SUPPLY_TYPES:
                errors.append(f"Supply Type '{v}' invalid (Inward/Outward)")

            # [1] Sub Type * (Required)
            v = val(1)
            if not v:
                errors.append("Sub Type is required")
            elif v not in self.VALID_SUB_TYPES:
                errors.append(f"Sub Type '{v}' invalid")

            # [2] Doc Type * (Required)
            v = val(2)
            if not v:
                errors.append("Doc Type is required")
            elif v not in self.VALID_DOC_TYPES:
                errors.append(f"Doc Type '{v}' invalid")

            # [3] Doc No * (Required, max 16 chars)
            v = val(3)
            if not v:
                errors.append("Doc No is required")
            elif len(v) > 16:
                errors.append(f"Doc No too long ({len(v)} chars, max 16)")

            # [4] Doc Date * (Required, DD/MM/YYYY format)
            v = val(4)
            norm_date = normalize_date(v)
            while len(row) < 5:
                row.append("")
            row[4] = norm_date
            v = norm_date

            m = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', v)
            if not m:
                errors.append(f"Doc Date '{v}' not in DD/MM/YYYY format")
            else:
                dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if mm < 1 or mm > 12 or dd < 1 or dd > 31:
                    errors.append(f"Doc Date '{v}' has invalid day/month")

            # [5] Transaction Type * (Required)
            v = val(5)
            if not v:
                errors.append("Transaction Type is required")
            elif v not in self.VALID_TRANSACTION_TYPES:
                errors.append(f"Transaction Type '{v}' invalid")

            # [7] From GSTIN * (Required, 15 chars, format check)
            v = val(7)
            if not v:
                errors.append("From GSTIN is required")
            elif v != "URP" and (len(v) != 15 or not re.match(r'^\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9]$', v)):
                errors.append(f"From GSTIN '{v}' invalid format")

            # [8] From_Address1 and [9] From_Address2 (Sanitize address)
            for col_idx in (8, 9):
                addr_v = val(col_idx)
                if addr_v:
                    cleaned_addr = sanitize_address(addr_v)
                    while len(row) < col_idx + 1:
                        row.append("")
                    row[col_idx] = cleaned_addr

            # [11] Dispatch Pin Code * (Required, 6 digits)
            v = val(11)
            if not v:
                errors.append("Dispatch Pin Code is required")
            elif not re.match(r'^\d{6}$', str(v).split('.')[0]):
                errors.append(f"Dispatch Pin Code '{v}' must be 6 digits")

            # [12] Bill From State * (Required)
            v = val(12)
            if not v:
                errors.append("Bill From State is required")
            else:
                norm_state = normalize_state(v)
                if not norm_state:
                    errors.append(f"Bill From State '{v}' not recognized")
                else:
                    while len(row) < 13:
                        row.append("")
                    row[12] = norm_state

            # [13] Dispatch From State *
            v = val(13)
            if not v:
                errors.append("Dispatch From State is required")
            else:
                norm_state = normalize_state(v)
                if not norm_state:
                    errors.append(f"Dispatch From State '{v}' not recognized")
                else:
                    while len(row) < 14:
                        row.append("")
                    row[13] = norm_state

            # [14] To_OtherPartyName (Required, sanitize to letters and spaces only)
            v = val(14)
            if not v:
                errors.append("To_OtherPartyName is required")
            else:
                cleaned = re.sub(r"[^a-zA-Z\s]", "", v)
                cleaned = re.sub(r"\s+", " ", cleaned).strip()
                while len(row) < 15:
                    row.append("")
                row[14] = cleaned

            # [15] To GSTIN * (Required)
            v = val(15)
            if not v:
                errors.append("To GSTIN is required")
            elif v != "URP" and (len(v) != 15 or not re.match(r'^\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9]$', v)):
                errors.append(f"To GSTIN '{v}' invalid format")

            # [16] To_Address1 and [17] To_Address2 (Sanitize address)
            for col_idx in (16, 17):
                addr_v = val(col_idx)
                if addr_v:
                    cleaned_addr = sanitize_address(addr_v)
                    while len(row) < col_idx + 1:
                        row.append("")
                    row[col_idx] = cleaned_addr

            # [19] Ship To Pin Code * (Required, 6 digits)
            v = val(19)
            if not v:
                errors.append("Ship To Pin Code is required")
            elif not re.match(r'^\d{6}$', str(v).split('.')[0]):
                errors.append(f"Ship To Pin Code '{v}' must be 6 digits")

            # [20] Bill To State * (Required)
            v = val(20)
            if not v:
                errors.append("Bill To State is required")
            else:
                norm_state = normalize_state(v)
                if not norm_state:
                    errors.append(f"Bill To State '{v}' not recognized")
                else:
                    while len(row) < 21:
                        row.append("")
                    row[20] = norm_state

            # [21] Ship To State * (Required)
            v = val(21)
            if not v:
                errors.append("Ship To State is required")
            else:
                norm_state = normalize_state(v)
                if not norm_state:
                    errors.append(f"Ship To State '{v}' not recognized")
                else:
                    while len(row) < 22:
                        row.append("")
                    row[21] = norm_state

            # [23] Description (Sanitize description text)
            v = val(23)
            if v:
                cleaned_desc = sanitize_address(v)
                while len(row) < 24:
                    row.append("")
                row[23] = cleaned_desc

            # [24] HSN * (Required, 4-8 digits)
            v = val(24)
            if not v:
                errors.append("HSN is required")
            else:
                hsn_str = str(v).split('.')[0]
                if not re.match(r'^\d{4,8}$', hsn_str):
                    errors.append(f"HSN '{v}' must be 4-8 digits")

            # [26] Qty (if provided, must be numeric > 0)
            v = val(26)
            if v:
                try:
                    qty = float(v)
                    if qty <= 0:
                        errors.append(f"Qty must be > 0 (got {v})")
                except ValueError:
                    errors.append(f"Qty '{v}' is not a valid number")

            # [27] Assessable Value * (Required, numeric)
            v = val(27)
            if not v:
                errors.append("Assessable Value is required")
            else:
                try:
                    float(v)
                except ValueError:
                    errors.append(f"Assessable Value '{v}' is not numeric")

            # [28] Tax Rate (format: S+C+I+Cess+CessNonAdvol)
            v = val(28)
            if v:
                parts = v.split('+')
                if len(parts) != 5:
                    errors.append(f"Tax Rate '{v}' must have 5 parts (S+C+I+Cess+CessNA)")
                else:
                    for p in parts:
                        try:
                            float(p)
                        except ValueError:
                            errors.append(f"Tax Rate part '{p}' is not numeric")
                            break

            # [35] Total Invoice Value * (Required, numeric)
            v = val(35)
            if not v:
                errors.append("Total Invoice Value is required")
            else:
                try:
                    float(v)
                except ValueError:
                    errors.append(f"Total Invoice Value '{v}' is not numeric")

            # --- Part B Validations ---

            # [36] Trans Mode (if provided, must be valid)
            v = val(36)
            if v and v not in self.VALID_TRANS_MODES:
                errors.append(f"Trans Mode '{v}' invalid (Road/Rail/Air/Ship)")

            # [37] Distance * (Required, 1-4000 km)
            v = val(37)
            if not v:
                errors.append("Distance is required")
            else:
                try:
                    dist = int(float(v))
                    if dist < 1 or dist > 4000:
                        errors.append(f"Distance {dist} out of range (1-4000 km)")
                except ValueError:
                    errors.append(f"Distance '{v}' is not numeric")

            # [42] Vehicle No (if Trans Mode is Road, should be present)
            trans_mode = val(36)
            vehicle = val(42)
            if trans_mode == "Road" and not vehicle:
                errors.append("Vehicle No required when Trans Mode is Road")

            # [43] Vehicle Type (if provided, must be valid)
            v = val(43)
            if v and v not in self.VALID_VEHICLE_TYPES:
                errors.append(f"Vehicle Type '{v}' invalid (Regular/Over Dimensional Cargo)")

            # Verify either Part-B details or Transporter details are present
            trans_mode_val = val(36)
            trans_id_val = val(39)
            vehicle_no_val = val(42)
            trans_doc_no_val = val(40)
            if not trans_id_val and not vehicle_no_val and not trans_doc_no_val:
                errors.append("Please enter either Part-B details (Vehicle No / Trans DocNo) or Transporter details (Trans ID)")

            # --- Write errors to Errors List column (index 45) ---
            if errors:
                total_errors += len(errors)
                error_str = "; ".join(errors)
                error_summary.append((row_num, errors))
                # Pad row if needed
                while len(row) < 46:
                    row.append("")
                row[45] = error_str
            else:
                while len(row) < 46:
                    row.append("")
                row[45] = "OK"

        # Refresh grid with error annotations
        self.sheet.set_sheet_data(grid_data)
        self.reapply_dropdowns()
        self.sheet.set_all_row_heights(height=20)
        self.sheet.refresh()

        # Show results
        if total_errors == 0:
            messagebox.showinfo(
                "Validation Passed ☑",
                f"All {len(grid_data)} rows passed validation.\n\n"
                f"The 'Errors List' column (last column) shows 'OK' for each row."
            )
            self.set_status(f"Validation passed. {len(grid_data)} rows are clean.")
        else:
            # Build summary text
            lines = []
            for row_num, errs in error_summary:
                lines.append(f"Row {row_num}: {len(errs)} error(s)")
                for e in errs:
                    lines.append(f"  • {e}")
                lines.append("") # Empty line between rows
            lines.append("Check the 'Errors List' column (last column) for details.")
            lines_text = "\n".join(lines)

            self.show_validation_errors(total_errors, len(error_summary), lines_text)
            self.set_status(f"Validation failed: {total_errors} errors in {len(error_summary)} rows.")

    # --- Formatting and Text Processing Helpers ---

    def split_address(self, addr):
        addr = addr.strip()
        if len(addr) <= 50:
            return addr, ""
        split_idx = 50
        for i in range(50, 0, -1):
            if addr[i] in [" ", ","]:
                split_idx = i
                break
        addr1 = addr[:split_idx].strip()
        addr2 = addr[split_idx:].strip()
        if addr2.startswith(","):
            addr2 = addr2[1:].strip()
        return addr1[:50], addr2[:50]

    def extract_pincode(self, addr):
        match = re.search(r"\b\d{5,6}\b", addr)
        if match:
            pin = match.group(0)
            if len(pin) == 5:
                pin += "1" # Correct standard truncated pincode
            return pin
        return ""

    def extract_place(self, addr, pincode):
        addr = addr.strip()
        parts = [p.strip() for p in addr.split(",") if p.strip()]
        if not parts:
            return "ERNAKULAM"
        for i, part in enumerate(parts):
            p_upper = part.upper()
            if "KERALA" in p_upper or (pincode and part == pincode):
                if i > 0:
                    return parts[i - 1][:50].upper()
        
        last_part = parts[-1]
        return last_part[:50].upper()

    def format_date(self, val):
        if val is None or val == "":
            return ""
        if isinstance(val, (int, float)):
            # Excel serial number date
            import datetime
            dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=val)
            return dt.strftime("%d/%m/%Y")
        
        str_val = str(val).strip()
        months = {"Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
                  "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"}
        
        # Format "05-Aug-2026"
        m = re.match(r"(\d{1,2})[-/ ]([A-Za-z]{3})[-/ ](\d{2,4})", str_val)
        if m:
            day = f"{int(m.group(1)):02d}"
            month = months.get(m.group(2).title(), "01")
            year = m.group(3)
            if len(year) == 2:
                year = "20" + year
            return f"{day}/{month}/{year}"

        # Format "05-08-2026"
        m_num = re.match(r"(\d{1,2})[-/ ](\d{1,2})[-/ ](\d{2,4})", str_val)
        if m_num:
            day = f"{int(m_num.group(1)):02d}"
            month = f"{int(m_num.group(2)):02d}"
            year = m_num.group(3)
            if len(year) == 2:
                year = "20" + year
            return f"{day}/{month}/{year}"

        return str_val

if __name__ == "__main__":
    app = EWayBillApp()
    app.mainloop()
